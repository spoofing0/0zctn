# -*- coding: utf-8 -*-
import re, json, os, asyncio, sys, pytz, logging
from datetime import datetime, timedelta
from telethon import TelegramClient, events
from telethon.errors import FloodWaitError, MessageNotModifiedError, RPCError
from telethon.tl.types import DocumentAttributeFilename
from collections import defaultdict, deque
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

# Logging ayarı
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('/var/log/royal_bot.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger('RoyalBot')

# Config değişkenleri
API_ID = 29581698
API_HASH = '0caabd4263f1d4e5f753659a787c2e7d'
BOT_TOKEN = 'YOUR_BOT_TOKEN_HERE'  # 🔑 BURAYA BOT TOKEN'INI YAZ!
KANAL_KAYNAK_ID = -1001626824569
KANAL_HEDEF = "@royalbaccfree"
ADMIN_ID = 1136442929
SISTEM_MODU = "normal_hibrit"
GMT3 = pytz.timezone('Europe/Istanbul')

# DEBUG MOD
DEBUG = True

# CLIENT TANIMI
client = TelegramClient('/root/0zctn/royal_bot_session', API_ID, API_HASH)

# Global değişkenler
game_results, martingale_trackers, color_trend, recent_games = {}, {}, [], []
MAX_MARTINGALE_STEPS, MAX_GAME_NUMBER, is_signal_active, daily_signal_count = 3, 1440, False, 0

# Excel dosyası
EXCEL_FILE = "/root/0zctn/royal_baccarat_data.xlsx"

# C2_3 Tipleri
C2_3_TYPES = {
    '#C2_3': {'emoji': '🔴', 'name': 'KLASİK', 'confidence': 0.9},
    '#C2_2': {'emoji': '🔵', 'name': 'ALTERNATİF', 'confidence': 0.7},
    '#C3_2': {'emoji': '🟢', 'name': 'VARYANT', 'confidence': 0.6},
    '#C3_3': {'emoji': '🟡', 'name': 'ÖZEL', 'confidence': 0.7}
}

# İstatistikler
performance_stats = {
    'total_signals': 0, 'win_signals': 0, 'loss_signals': 0, 'total_profit': 0,
    'current_streak': 0, 'max_streak': 0,
    'daily_stats': defaultdict(lambda: {'signals': 0, 'wins': 0, 'losses': 0, 'profit': 0}),
    'signal_history': deque(maxlen=1000)
}

# DEBUG FONKSİYONLARI
def debug_log(message):
    if DEBUG:
        logger.info(f"🔍 DEBUG: {message}")

# EXCEL FONKSİYONLARI
def init_excel_file():
    try:
        if not os.path.exists(EXCEL_FILE):
            wb = Workbook()
            ws = wb.active
            ws.title = "Baccarat Data"
            headers = ["Oyun No", "Tarih", "Saat", "Player Kartları", "Banker Kartları", "Kazanan", "C2 Tipi", "Renk Tahmini", "Pattern Tipi", "Sonuç", "Kazanç/Kayıp", "Martingale Seviye"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header).font = Font(bold=True)
            wb.save(EXCEL_FILE)
            debug_log("✅ Excel dosyası oluşturuldu")
    except Exception as e:
        debug_log(f"❌ Excel oluşturma hatası: {e}")

def save_to_excel(game_data):
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        row = ws.max_row + 1
        
        for col, key in enumerate([
            'game_number', 'date', 'time', 'player_cards', 'banker_cards', 
            'winner', 'c2_type', 'color_prediction', 'pattern_type', 
            'result', 'profit_loss', 'martingale_level'
        ], 1):
            ws.cell(row=row, column=col, value=game_data.get(key))
            
        wb.save(EXCEL_FILE)
    except Exception as e:
        debug_log(f"❌ Excel kaydetme hatası: {e}")

async def save_game_result_to_excel(game_info):
    try:
        game_number = game_info['game_number']
        now = datetime.now(GMT3)
        
        game_data = {
            'game_number': game_number,
            'date': now.strftime('%d.%m.%Y'),
            'time': now.strftime('%H:%M:%S'),
            'player_cards': game_info['player_cards'],
            'banker_cards': game_info['banker_cards'],
            'winner': 'Tie',
            'c2_type': game_info.get('c2_3_type', 'N/A'),
            'color_prediction': 'N/A',
            'pattern_type': 'N/A',
            'result': 'N/A',
            'profit_loss': 0,
            'martingale_level': 0
        }
        
        save_to_excel(game_data)
        debug_log(f"✅ Oyun #{game_number} Excel'e kaydedildi")
        
    except Exception as e:
        debug_log(f"❌ Oyun kaydetme hatası: {e}")

# MARTINGALE SİSTEMİ - TAMAMEN YENİ
async def process_martingale_progression(tracker, result):
    """Martingale ilerlemesini işle"""
    try:
        current_step = tracker['step']
        signal_suit = tracker['signal_suit']
        original_game_num = tracker['sent_game_number']
        
        if result == "KAYIP" and current_step < MAX_MARTINGALE_STEPS - 1:
            # Martingale devam et
            next_step = current_step + 1
            next_game_num = get_next_game_number(tracker['expected_game_number_for_check'])
            
            debug_log(f"🔄 Martingale ilerliyor: {current_step} -> {next_step}. seviye")
            
            # Yeni martingale sinyali gönder
            await send_martingale_signal(
                game_num=next_game_num,
                signal_suit=signal_suit,
                reason=tracker['reason'],
                c2_3_info={'c2_3_type': tracker['c2_3_type'], 'c2_3_description': tracker['c2_3_description']},
                step=next_step,
                original_game_num=original_game_num
            )
            
            return True  # Martingale devam ediyor
        else:
            # Martingale tamamlandı
            debug_log(f"✅ Martingale tamamlandı: {current_step}. seviye")
            return False  # Martingale bitti
            
    except Exception as e:
        debug_log(f"❌ Martingale işleme hatası: {e}")
        return False

async def send_martingale_signal(game_num, signal_suit, reason, c2_3_info=None, step=0, original_game_num=None):
    """Martingale sinyali gönder"""
    global is_signal_active
    
    try:
        suit_display = get_suit_display_name(signal_suit)
        
        if c2_3_info:
            c2_3_type = c2_3_info.get('c2_3_type', '#C2_3')
            c2_3_desc = c2_3_info.get('c2_3_description', 'KLASİK')
            trigger_info = f"{c2_3_desc} {c2_3_type}"
        else: 
            trigger_info = "KLASİK #C2_3"
        
        gmt3_time = datetime.now(GMT3).strftime('%H:%M:%S')
        
        if step == 0:
            signal_text = f"""🎯 **SİNYAL BAŞLADI** 🎯
#N{game_num} - {suit_display}
📊 Tetikleyici: {trigger_info}
🎯 Sebep: {reason}
⚡ Strateji: Martingale {MAX_MARTINGALE_STEPS} Seviye
🕒 {gmt3_time} (GMT+3)
🔴 SONUÇ: BEKLENİYOR..."""
        else:
            signal_text = f"""🔄 **MARTINGALE SİNYALİ** 🔄
#N{game_num} - {suit_display}
📊 Önceki Oyun: #N{original_game_num}
🎯 Seviye: {step}. Seviye
🎯 Sebep: {reason}
⚡ Strateji: Martingale {MAX_MARTINGALE_STEPS} Seviye
🕒 {gmt3_time} (GMT+3)
🔴 SONUÇ: BEKLENİYOR..."""
        
        debug_log(f"📤 Martingale sinyali gönderiliyor: Oyun #{game_num} - Seviye {step}")
        sent_message = await client.send_message(KANAL_HEDEF, signal_text)
        
        if step == 0:
            daily_signal_count += 1
        
        martingale_trackers[game_num] = {
            'message_obj': sent_message, 
            'step': step, 
            'signal_suit': signal_suit, 
            'sent_game_number': game_num, 
            'expected_game_number_for_check': game_num,
            'original_game_num': original_game_num or game_num,
            'start_time': datetime.now(GMT3), 
            'reason': reason, 
            'c2_3_type': c2_3_info.get('c2_3_type', '#C2_3') if c2_3_info else '#C2_3',
            'c2_3_description': c2_3_info.get('c2_3_description', 'KLASİK') if c2_3_info else 'KLASİK',
            'results': []
        }
        is_signal_active = True
        debug_log(f"✅ Martingale sinyali gönderildi: #{game_num} - Seviye {step}")
        
    except Exception as e: 
        debug_log(f"❌ Martingale sinyal gönderme hatası: {e}")

# SONUÇ TAKİP SİSTEMİ - GÜNCELLENDİ
async def check_and_update_signal_results():
    """Aktif sinyallerin sonuçlarını kontrol et ve güncelle"""
    global is_signal_active, martingale_trackers
    
    try:
        current_game_numbers = list(game_results.keys())
        debug_log(f"🔍 Sonuç kontrolü: {len(martingale_trackers)} sinyal, {len(current_game_numbers)} oyun")
        
        for game_num, tracker in list(martingale_trackers.items()):
            expected_game = tracker['expected_game_number_for_check']
            debug_log(f"🔍 Sinyal #{game_num} için beklenen oyun: #{expected_game}")
            
            # Beklenen oyun numarası mevcut mu?
            if expected_game in game_results:
                debug_log(f"✅ Beklenen oyun #{expected_game} bulundu, sonuç güncelleniyor...")
                game_info = game_results[expected_game]
                signal_suit = tracker['signal_suit']
                
                # Kazanç/kayıp kontrolü
                actual_suit = extract_largest_value_suit(game_info['player_cards'])
                debug_log(f"🎯 Sinyal rengi: {signal_suit}, Gerçek renk: {actual_suit}")
                
                # Renk karşılaştırması
                if actual_suit == signal_suit:
                    result = "KAZANÇ"
                    result_emoji = "🟢"
                    result_text = "KAZANDINIZ! 🎉"
                else:
                    result = "KAYIP" 
                    result_emoji = "🔴"
                    result_text = "KAYBETTİNİZ! 💸"
                
                debug_log(f"🎲 Sonuç: {result}")
                
                # Mesajı güncelle
                try:
                    duration = datetime.now(GMT3) - tracker['start_time']
                    duration_str = f"{duration.seconds}s"
                    
                    if tracker['step'] == 0:
                        new_text = f"""✅ {result} ✅
#N{game_num} - {get_suit_display_name(signal_suit)}
📊 Sebep: {tracker['reason']}
🎯 Seviye: {tracker['step']}. Seviye
⏱️ Süre: {duration_str}
🕒 Bitiş: {datetime.now(GMT3).strftime('%H:%M:%S')}
🏆 SONUÇ: {result_text}"""
                    else:
                        new_text = f"""🔄 {result} 🔄
#N{game_num} - {get_suit_display_name(signal_suit)}
📊 Önceki Oyun: #N{tracker['original_game_num']}
🎯 Seviye: {tracker['step']}. Seviye
⏱️ Süre: {duration_str}
🕒 Bitiş: {datetime.now(GMT3).strftime('%H:%M:%S')}
🏆 SONUÇ: {result_text}"""
                    
                    await tracker['message_obj'].edit(new_text)
                    debug_log(f"✅ Sinyal mesajı güncellendi: {result}")
                except Exception as e:
                    debug_log(f"❌ Mesaj güncelleme hatası: {e}")
                
                # Excel'de sonucu güncelle
                await update_excel_with_result(expected_game, result, tracker['step'])
                
                # İstatistikleri güncelle
                update_performance_stats(result, tracker)
                
                # Martingale ilerlemesini kontrol et
                martingale_continues = await process_martingale_progression(tracker, result)
                
                # Eğer martingale devam etmiyorsa tracker'ı sil
                if not martingale_continues:
                    del martingale_trackers[game_num]
                    debug_log(f"✅ Sinyal #{game_num} sonuçlandı: {result}")
                else:
                    debug_log(f"🔄 Martingale devam ediyor: #{game_num}")
            else:
                debug_log(f"⏳ Beklenen oyun #{expected_game} henüz yok")
        
        # Aktif sinyal kalmadıysa durumu güncelle
        if not martingale_trackers:
            is_signal_active = False
            debug_log("🔔 Aktif sinyal kalmadı")
            
    except Exception as e:
        debug_log(f"❌ Sonuç kontrol hatası: {e}")

async def update_excel_with_result(game_number, result, martingale_level):
    """Excel'deki sonuç sütununu güncelle"""
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == game_number:
                # Sonuç sütununu güncelle
                ws.cell(row=row, column=10, value=result)
                
                # Kazanç/Kayıp sütununu güncelle
                profit_loss = 1 if result == "KAZANÇ" else -1
                ws.cell(row=row, column=11, value=profit_loss)
                
                # Martingale seviyesini güncelle
                ws.cell(row=row, column=12, value=martingale_level)
                
                break
                
        wb.save(EXCEL_FILE)
        debug_log(f"✅ Excel'de #{game_number} sonucu güncellendi: {result} (Seviye: {martingale_level})")
        
    except Exception as e:
        debug_log(f"❌ Excel sonuç güncelleme hatası: {e}")

def update_performance_stats(result, tracker):
    """Performans istatistiklerini güncelle"""
    try:
        if tracker['step'] == 0:  # Sadece ilk sinyali say
            performance_stats['total_signals'] += 1
            
            if result == "KAZANÇ":
                performance_stats['win_signals'] += 1
                performance_stats['current_streak'] += 1
                performance_stats['total_profit'] += 1
            else:
                performance_stats['loss_signals'] += 1
                performance_stats['current_streak'] = 0
                performance_stats['total_profit'] -= 1
                
            performance_stats['max_streak'] = max(
                performance_stats['max_streak'], 
                performance_stats['current_streak']
            )
            
            # Günlük istatistikleri güncelle
            today = datetime.now(GMT3).strftime('%Y-%m-%d')
            performance_stats['daily_stats'][today]['signals'] += 1
            if result == "KAZANÇ":
                performance_stats['daily_stats'][today]['wins'] += 1
                performance_stats['daily_stats'][today]['profit'] += 1
            else:
                performance_stats['daily_stats'][today]['losses'] += 1
                performance_stats['daily_stats'][today]['profit'] -= 1
        
    except Exception as e:
        debug_log(f"❌ İstatistik güncelleme hatası: {e}")

# TEMEL FONKSİYONLAR
def get_suit_display_name(suit_symbol):
    suit_names = {'♠': '♠️ MAÇA', '♥': '❤️ KALP', '♦': '♦️ KARO', '♣': '♣️ SİNEK'}
    return suit_names.get(suit_symbol, f"❓ {suit_symbol}")

def get_baccarat_value(card_char):
    if card_char == '10': return 10
    if card_char in 'AKQJ2T': return 0
    elif card_char.isdigit(): return int(card_char)
    return -1

def extract_largest_value_suit(cards_str):
    try:
        cards = re.findall(r'(10|[A2-9TJQK])([♣♦♥♠])', cards_str)
        if not cards: return None
        max_value, largest_value_suit = -1, None
        for card_char, suit in cards:
            value = get_baccarat_value(card_char)
            if value > max_value: 
                max_value, largest_value_suit = value, suit
        return largest_value_suit if max_value > 0 else None
    except Exception as e:
        debug_log(f"❌ extract_largest_value_suit hatası: {e}")
        return None

def analyze_simple_pattern(player_cards, banker_cards, game_number):
    try:
        signal_color = extract_largest_value_suit(player_cards)
        if not signal_color: 
            return None, "Renk tespit edilemedi"
        
        color_trend.append(signal_color)
        if len(color_trend) > 10: 
            color_trend.pop(0)
            
        # Basitleştirilmiş pattern analizi
        player_card_data = re.findall(r'(10|[A2-9TJQK])([♣♦♥♠])', player_cards)
        player_values = [get_baccarat_value(card[0]) for card in player_card_data]
        
        if sum(player_values) in [8, 9]:
            return signal_color, "🏆 DOĞAL KAZANÇ"
        elif sum(player_values) >= 8 and len(player_values) >= 3:
            return signal_color, "🎯 GÜÇLÜ EL"
        else:
            return signal_color, "📈 STANDART SİNYAL"
            
    except Exception as e:
        return None, f"Hata: {e}"

def get_next_game_number(current_game_num):
    next_num = current_game_num + 1
    return 1 if next_num > MAX_GAME_NUMBER else next_num

# SİNYAL SİSTEMLERİ
async def send_new_signal(game_num, signal_suit, reason, c2_3_info=None):
    """Yeni sinyal gönder - Martingale başlangıcı"""
    await send_martingale_signal(game_num, signal_suit, reason, c2_3_info, step=0)

async def normal_hibrit_sistemi(game_info):
    try:
        trigger_game_num = game_info['game_number']
        c2_3_info = {
            'c2_3_type': game_info.get('c2_3_type'), 
            'c2_3_description': game_info.get('c2_3_description')
        }
        
        signal_color, reason = analyze_simple_pattern(
            game_info['player_cards'], 
            game_info['banker_cards'], 
            trigger_game_num
        )
        
        if signal_color:
            # BİR SONRAKİ OYUN İÇİN SİNYAL GÖNDER
            next_game_num = get_next_game_number(trigger_game_num)
            debug_log(f"🔄 Sinyal tetiklendi: #{trigger_game_num} -> #{next_game_num}")
            await send_new_signal(next_game_num, signal_color, reason, c2_3_info)
        else:
            debug_log(f"❌ Sinyal tetiklenemedi: Renk bulunamadı")
            
    except Exception as e:
        debug_log(f"❌ Normal hibrit sistemi hatası: {e}")

async def excel_based_sistemi(game_info):
    try:
        debug_log("📊 Excel sistemi aktif")
        await normal_hibrit_sistemi(game_info)
    except Exception as e: 
        debug_log(f"❌ Excel sistemi hatası: {e}")

# KOMUTLAR
@client.on(events.NewMessage(pattern='(?i)/basla'))
async def handle_start(event): 
    await event.reply("🤖 Royal Baccarat Bot Aktif! 🎯")
    debug_log("✅ /basla komutu çalıştı")

@client.on(events.NewMessage(pattern='(?i)/durum'))
async def handle_durum(event):
    try:
        me = await client.get_me()
        bot_username = f"@{me.username}"
    except:
        bot_username = "Bilinmiyor"
        
    aktif_sinyal = "✅ Evet" if is_signal_active else "❌ Hayır"
    gmt3_time = datetime.now(GMT3).strftime('%H:%M:%S')
    
    total = performance_stats['total_signals']
    wins = performance_stats['win_signals']
    win_rate = (wins / total * 100) if total > 0 else 0
    
    await event.reply(f"""🤖 **ROYAL BACCARAT BOT** 🎯

🤖 **Bot:** {bot_username}
🟢 **Durum:** Çalışıyor
🎯 **Aktif Sinyal:** {aktif_sinyal} ({len(martingale_trackers)} sinyal)
🎛️ **Mod:** {SISTEM_MODU}
🕒 **Saat:** {gmt3_time} (GMT+3)
📨 **Günlük Sinyal:** {daily_signal_count}

📊 **İSTATİSTİKLER:**
• Toplam Sinyal: {total}
• Kazanç: {wins} | Kayıp: {performance_stats['loss_signals']}
• Başarı Oranı: %{win_rate:.1f}
• Toplam Kâr: {performance_stats['total_profit']}
• Seri: {performance_stats['current_streak']} (Rekor: {performance_stats['max_streak']})

⚡ **Martingale:** {MAX_MARTINGALE_STEPS} Seviye Aktif""")

@client.on(events.NewMessage(pattern='(?i)/test'))
async def handle_test(event):
    """Test sinyali gönder"""
    try:
        debug_log("🧪 Test sinyali gönderiliyor...")
        test_game_num = 999
        test_suit = '♠'
        test_reason = "TEST SİNYALİ"
        
        await send_new_signal(test_game_num, test_suit, test_reason)
        await event.reply("✅ Test sinyali gönderildi! Hedef kanalı kontrol et.")
    except Exception as e:
        await event.reply(f"❌ Test hatası: {e}")

@client.on(events.NewMessage(pattern='(?i)/debug'))
async def handle_debug(event):
    """Debug bilgileri"""
    active_signals = []
    for game_num, tracker in martingale_trackers.items():
        active_signals.append(f"#{game_num} - {tracker['signal_suit']} (Seviye: {tracker['step']})")
    
    debug_info = f"""🔧 **DEBUG BİLGİLERİ**

📊 Game Results: {len(game_results)} oyun
🎯 Active Signals: {len(martingale_trackers)}
📈 Color Trend: {color_trend[-5:] if color_trend else 'Boş'}
🔔 Signal Active: {is_signal_active}
📨 Daily Signals: {daily_signal_count}

🔍 **Aktif Sinyaller:**
{chr(10).join(active_signals) if active_signals else 'Aktif sinyal yok'}

⚡ **Martingale Sistem:**
• Maksimum Seviye: {MAX_MARTINGALE_STEPS}
• Aktif Trackers: {len(martingale_trackers)}"""

    await event.reply(debug_info)

# ANA MESAJ İŞLEYİCİ
@client.on(events.NewMessage(chats=KANAL_KAYNAK_ID))
async def handle_source_channel_message(event):
    try:
        text = event.message.text or ""
        debug_log(f"📥 Kaynak kanaldan mesaj: {text[:100]}...")
        
        # GELİŞMİŞ PARSE YÖNTEMİ
        game_match = re.search(r'#N(\d+)', text)
        player_match = re.search(r'[Pp]layer[^()]*\(([^)]+)\)', text)
        banker_match = re.search(r'[Bb]anker[^()]*\(([^)]+)\)', text)
        
        debug_log(f"🔍 Parse sonuçları - Game: {game_match}, Player: {player_match}, Banker: {banker_match}")
        
        if game_match and player_match:
            game_info = {
                'game_number': int(game_match.group(1)),
                'player_cards': player_match.group(1),
                'banker_cards': banker_match.group(1) if banker_match else '',
                'is_final': True,
                'is_c2_3': any(trigger in text for trigger in C2_3_TYPES.keys()),
                'c2_3_type': next((trigger for trigger in C2_3_TYPES.keys() if trigger in text), '#C2_3'),
                'c2_3_description': C2_3_TYPES.get(next((trigger for trigger in C2_3_TYPES.keys() if trigger in text), '#C2_3'), {}).get('name', 'KLASİK')
            }
            
            debug_log(f"🎮 Oyun bilgisi: #{game_info['game_number']}, C2: {game_info['is_c2_3']}")
            
            # OYUNU KAYDET
            game_results[game_info['game_number']] = game_info
            await save_game_result_to_excel(game_info)
            
            # SONUÇ KONTROLÜNÜ YAP
            debug_log("🔄 Sonuç kontrolü yapılıyor...")
            await check_and_update_signal_results()
            debug_log("✅ Sonuç kontrolü tamamlandı")
            
            # YENİ SİNYAL TETİKLEME
            sinyal_kosul = not is_signal_active and game_info['is_final'] and game_info.get('is_c2_3')
            debug_log(f"🔔 Sinyal koşulu: {sinyal_kosul} (active={is_signal_active}, final={game_info['is_final']}, c2_3={game_info.get('is_c2_3')})")
            
            if sinyal_kosul:
                debug_log("🎯 Sinyal tetikleniyor...")
                if SISTEM_MODU == "normal_hibrit": 
                    await normal_hibrit_sistemi(game_info)
                elif SISTEM_MODU == "excel_based":
                    await excel_based_sistemi(game_info)
            else:
                debug_log("⏸️ Sinyal tetiklenmedi - Koşullar sağlanmıyor")
                    
    except Exception as e: 
        debug_log(f"❌ Mesaj işleme hatası: {e}")

# ANA ÇALIŞTIRMA
async def main():
    try:
        debug_log("🚀 Bot başlatılıyor...")
        await client.start(bot_token=BOT_TOKEN)
        
        me = await client.get_me()
        debug_log(f"✅ Bot başarıyla başlatıldı: @{me.username}")
        
        await client.run_until_disconnected()
        
    except Exception as e:
        debug_log(f"❌ Bot başlatma hatası: {e}")

if __name__ == '__main__':
    print("🤖 ROYAL BACCARAT BOT BAŞLIYOR...")
    print("🔍 DEBUG MOD: AKTİF")
    print("⚡ MARTINGALE SİSTEMİ: AKTİF")
    
    # Eski session dosyalarını temizle
    session_files = ['/root/0zctn/royal_bot_session.session', 'royal_bot.session']
    for session_file in session_files:
        if os.path.exists(session_file):
            try:
                os.remove(session_file)
                print(f"✅ Eski session silindi: {session_file}")
            except Exception as e:
                print(f"⚠️ Session silinemedi: {e}")
    
    init_excel_file()
    
    try:
        if BOT_TOKEN == 'YOUR_BOT_TOKEN_HERE' or not BOT_TOKEN:
            print("❌ ❌ ❌ HATA: BOT_TOKEN ayarlanmamış!")
            print("🔧 Çözüm: Bot tokenını BOT_TOKEN değişkenine yaz")
            sys.exit(1)
            
        client.loop.run_until_complete(main())
    except KeyboardInterrupt:
        print("🔴 Bot kullanıcı tarafından durduruldu")
    except Exception as e:
        print(f"❌ Bot hatası: {e}")
    finally:
        print("🔴 Bot durduruldu")