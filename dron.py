# -*- coding: utf-8 -*-
import re, json, os, asyncio, sys, pytz, logging
from datetime import datetime, timedelta
from telethon import TelegramClient, events
from telethon.errors import FloodWaitError, MessageNotModifiedError, RPCError
from telethon.tl.types import DocumentAttributeFilename
from collections import defaultdict, deque
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

# Logging ayarı
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('/var/log/dron_bot.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger('DronBot')

# Config değişkenleri
API_ID = 29581698
API_HASH = '0caabd4263f1d4e5f753659a787c2e7d'
BOT_TOKEN = ''  # 🔑 BURAYA TOKEN'INI YAZ
KANAL_KAYNAK_ID = -1001626824569
KANAL_HEDEF = "@royalbaccfree"
ADMIN_ID = 1136442929
SISTEM_MODU = "normal_hibrit"
GMT3 = pytz.timezone('Europe/Istanbul')

# CLIENT TANIMI
client = TelegramClient('/root/0zctn/dron_bot_session', API_ID, API_HASH)

# Global değişkenler
game_results, martingale_trackers, color_trend = {}, {}, []
MAX_MARTINGALE_STEPS, MAX_GAME_NUMBER, is_signal_active, daily_signal_count = 3, 1440, False, 0

# Excel dosyası
EXCEL_FILE = "/root/0zctn/dron_baccarat_data.xlsx"
RED_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
GREEN_FILL = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

# İstatistikler
performance_stats = {
    'total_signals': 0, 'win_signals': 0, 'loss_signals': 0, 'total_profit': 0,
    'current_streak': 0, 'max_streak': 0,
    'daily_stats': defaultdict(lambda: {'signals': 0, 'wins': 0, 'losses': 0, 'profit': 0}),
    'signal_history': deque(maxlen=1000)
}

# DEBUG FONKSİYONLARI
def debug_log(message):
    logger.info(f"🔍 {message}")

# EXCEL FONKSİYONLARI
def init_excel_file():
    try:
        if not os.path.exists(EXCEL_FILE):
            wb = Workbook()
            ws = wb.active
            ws.title = "Baccarat Data"
            headers = ["Oyun No", "Tarih", "Saat", "Player Kartları", "Banker Kartları", "Kazanan", "Renk Tahmini", "Pattern Tipi", "Sonuç", "Kazanç/Kayıp", "Martingale Seviye"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header).font = Font(bold=True)
            wb.save(EXCEL_FILE)
            debug_log("✅ Excel dosyası oluşturuldu")
    except Exception as e:
        debug_log(f"❌ Excel oluşturma hatası: {e}")

def save_to_excel(game_data):
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        row = ws.max_row + 1
        
        ws.cell(row=row, column=1, value=game_data.get('game_number'))
        ws.cell(row=row, column=2, value=game_data.get('date'))
        ws.cell(row=row, column=3, value=game_data.get('time'))
        ws.cell(row=row, column=4, value=game_data.get('player_cards'))
        ws.cell(row=row, column=5, value=game_data.get('banker_cards'))
        ws.cell(row=row, column=6, value=game_data.get('winner'))
        ws.cell(row=row, column=7, value=game_data.get('color_prediction'))
        ws.cell(row=row, column=8, value=game_data.get('pattern_type'))
        
        result_cell = ws.cell(row=row, column=9, value=game_data.get('result'))
        if game_data.get('result') == 'KAZANÇ': 
            result_cell.fill = GREEN_FILL
        elif game_data.get('result') == 'KAYIP': 
            result_cell.fill = RED_FILL
            
        ws.cell(row=row, column=10, value=game_data.get('profit_loss'))
        ws.cell(row=row, column=11, value=game_data.get('martingale_level'))
        
        wb.save(EXCEL_FILE)
    except Exception as e:
        debug_log(f"❌ Excel kaydetme hatası: {e}")

# TEMEL FONKSİYONLAR
def get_suit_display_name(suit_symbol):
    suit_names = {'♠': '♠️ MAÇA', '♥': '❤️ KALP', '♦': '♦️ KARO', '♣': '♣️ SİNEK'}
    return suit_names.get(suit_symbol, f"❓ {suit_symbol}")

def extract_largest_value_suit(cards_str):
    try:
        cards = re.findall(r'([A2-9TJQK])([♣♦♥♠])', cards_str)
        if not cards: 
            return '♠'
        return cards[0][1]  # İlk kartın rengini al
    except Exception as e:
        debug_log(f"❌ Renk çıkarma hatası: {e}")
        return '♠'

def analyze_simple_pattern(player_cards, banker_cards, game_number):
    try:
        signal_color = extract_largest_value_suit(player_cards)
        if not signal_color: 
            return None, "Renk tespit edilemedi"
        
        return signal_color, "🎯 OTOMATİK SİNYAL"
    except Exception as e:
        return None, f"Hata: {e}"

def get_next_game_number(current_game_num):
    next_num = current_game_num + 1
    return 1 if next_num > MAX_GAME_NUMBER else next_num

# MARTINGALE SİSTEMİ
async def process_martingale_progression(tracker, result):
    """Martingale ilerlemesini işle"""
    try:
        current_step = tracker['step']
        signal_suit = tracker['signal_suit']
        original_game_num = tracker['original_game_num']
        
        if result == "KAYIP" and current_step < MAX_MARTINGALE_STEPS - 1:
            next_step = current_step + 1
            next_game_num = get_next_game_number(tracker['expected_game_number_for_check'])
            
            debug_log(f"🔄 Martingale ilerliyor: {current_step} -> {next_step}. seviye")
            
            await send_martingale_signal(
                game_num=next_game_num,
                signal_suit=signal_suit,
                reason=tracker['reason'],
                step=next_step,
                original_game_num=original_game_num
            )
            return True
        else:
            debug_log(f"✅ Martingale tamamlandı: {current_step}. seviye")
            return False
    except Exception as e:
        debug_log(f"❌ Martingale işleme hatası: {e}")
        return False

async def send_martingale_signal(game_num, signal_suit, reason, step=0, original_game_num=None):
    """Martingale sinyali gönder"""
    global is_signal_active, daily_signal_count
    
    try:
        suit_display = get_suit_display_name(signal_suit)
        gmt3_time = datetime.now(GMT3).strftime('%H:%M:%S')
        
        if step == 0:
            signal_text = f"""🎯 **SİNYAL BAŞLADI** 🎯
#N{game_num} - {suit_display}
📊 Sebep: {reason}
⚡ Strateji: Martingale {MAX_MARTINGALE_STEPS} Seviye
🕒 {gmt3_time} (GMT+3)
🔴 SONUÇ: BEKLENİYOR..."""
            daily_signal_count += 1
        else:
            signal_text = f"""🔄 **MARTINGALE SİNYALİ** 🔄
#N{game_num} - {suit_display}
📊 Önceki Oyun: #N{original_game_num}
🎯 Seviye: {step}. Seviye
⚡ Strateji: Martingale {MAX_MARTINGALE_STEPS} Seviye
🕒 {gmt3_time} (GMT+3)
🔴 SONUÇ: BEKLENİYOR..."""
        
        debug_log(f"📤 Sinyal gönderiliyor: Oyun #{game_num} - Seviye {step}")
        sent_message = await client.send_message(KANAL_HEDEF, signal_text)
        
        martingale_trackers[game_num] = {
            'message_obj': sent_message, 
            'step': step, 
            'signal_suit': signal_suit, 
            'sent_game_number': game_num, 
            'expected_game_number_for_check': game_num,
            'original_game_num': original_game_num or game_num,
            'start_time': datetime.now(GMT3), 
            'reason': reason
        }
        is_signal_active = True
        debug_log(f"✅ Sinyal gönderildi: #{game_num} - Seviye {step}")
        
    except Exception as e: 
        debug_log(f"❌ Sinyal gönderme hatası: {e}")

# SONUÇ TAKİP SİSTEMİ
async def check_and_update_signal_results():
    """Aktif sinyallerin sonuçlarını kontrol et ve güncelle"""
    global is_signal_active, martingale_trackers
    
    try:
        for game_num, tracker in list(martingale_trackers.items()):
            expected_game = tracker['expected_game_number_for_check']
            
            if expected_game in game_results:
                debug_log(f"✅ Beklenen oyun #{expected_game} bulundu")
                game_info = game_results[expected_game]
                signal_suit = tracker['signal_suit']
                
                # Kazanç/kayıp kontrolü
                actual_suit = extract_largest_value_suit(game_info['player_cards'])
                debug_log(f"🎯 Sinyal: {signal_suit}, Gerçek: {actual_suit}")
                
                if actual_suit == signal_suit:
                    result = "KAZANÇ"
                    result_emoji = "🟢"
                    result_text = "KAZANDINIZ! 🎉"
                    profit = 1
                else:
                    result = "KAYIP" 
                    result_emoji = "🔴"
                    result_text = "KAYBETTİNİZ! 💸"
                    profit = -1
                
                debug_log(f"🎲 Sonuç: {result}")
                
                # Mesajı güncelle
                try:
                    duration = datetime.now(GMT3) - tracker['start_time']
                    duration_str = f"{duration.seconds}s"
                    
                    if tracker['step'] == 0:
                        new_text = f"""✅ {result} ✅
#N{game_num} - {get_suit_display_name(signal_suit)}
📊 Sebep: {tracker['reason']}
🎯 Seviye: {tracker['step']}. Seviye
⏱️ Süre: {duration_str}
🕒 Bitiş: {datetime.now(GMT3).strftime('%H:%M:%S')}
🏆 SONUÇ: {result_text}"""
                    else:
                        new_text = f"""🔄 {result} 🔄
#N{game_num} - {get_suit_display_name(signal_suit)}
📊 Önceki Oyun: #N{tracker['original_game_num']}
🎯 Seviye: {tracker['step']}. Seviye
⏱️ Süre: {duration_str}
🕒 Bitiş: {datetime.now(GMT3).strftime('%H:%M:%S')}
🏆 SONUÇ: {result_text}"""
                    
                    await tracker['message_obj'].edit(new_text)
                    debug_log(f"✅ Mesaj güncellendi: {result}")
                except Exception as e:
                    debug_log(f"❌ Mesaj güncelleme hatası: {e}")
                
                # Excel'e kaydet
                await update_excel_with_result(expected_game, result, profit, tracker['step'])
                
                # İstatistikleri güncelle
                update_performance_stats(result, profit, tracker)
                
                # Martingale ilerlemesini kontrol et
                martingale_continues = await process_martingale_progression(tracker, result)
                
                if not martingale_continues:
                    del martingale_trackers[game_num]
                    debug_log(f"✅ Sinyal #{game_num} sonuçlandı")
        
        if not martingale_trackers:
            is_signal_active = False
            debug_log("🔔 Aktif sinyal kalmadı")
            
    except Exception as e:
        debug_log(f"❌ Sonuç kontrol hatası: {e}")

async def update_excel_with_result(game_number, result, profit, martingale_level):
    """Excel'deki sonuç sütununu güncelle"""
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == game_number:
                ws.cell(row=row, column=9, value=result)
                ws.cell(row=row, column=10, value=profit)
                ws.cell(row=row, column=11, value=martingale_level)
                break
                
        wb.save(EXCEL_FILE)
        debug_log(f"✅ Excel güncellendi: #{game_number} - {result}")
        
    except Exception as e:
        debug_log(f"❌ Excel güncelleme hatası: {e}")

def update_performance_stats(result, profit, tracker):
    """Performans istatistiklerini güncelle"""
    try:
        if tracker['step'] == 0:
            performance_stats['total_signals'] += 1
            
            if result == "KAZANÇ":
                performance_stats['win_signals'] += 1
                performance_stats['current_streak'] += 1
                performance_stats['total_profit'] += profit
            else:
                performance_stats['loss_signals'] += 1
                performance_stats['current_streak'] = 0
                performance_stats['total_profit'] += profit
                
            performance_stats['max_streak'] = max(
                performance_stats['max_streak'], 
                performance_stats['current_streak']
            )
            
            today = datetime.now(GMT3).strftime('%Y-%m-%d')
            performance_stats['daily_stats'][today]['signals'] += 1
            if result == "KAZANÇ":
                performance_stats['daily_stats'][today]['wins'] += 1
            else:
                performance_stats['daily_stats'][today]['losses'] += 1
            performance_stats['daily_stats'][today]['profit'] += profit
        
    except Exception as e:
        debug_log(f"❌ İstatistik güncelleme hatası: {e}")

# SİNYAL SİSTEMLERİ
async def send_new_signal(game_num, signal_suit, reason):
    """Yeni sinyal gönder"""
    await send_martingale_signal(game_num, signal_suit, reason, step=0)

async def normal_hibrit_sistemi(game_info):
    try:
        trigger_game_num = game_info['game_number']
        
        signal_color, reason = analyze_simple_pattern(
            game_info['player_cards'], 
            game_info['banker_cards'], 
            trigger_game_num
        )
        
        if signal_color:
            next_game_num = get_next_game_number(trigger_game_num)
            debug_log(f"🔄 Sinyal tetiklendi: #{trigger_game_num} -> #{next_game_num}")
            await send_new_signal(next_game_num, signal_color, reason)
        else:
            debug_log("❌ Sinyal tetiklenemedi")
            
    except Exception as e:
        debug_log(f"❌ Sistem hatası: {e}")

# KOMUTLAR
@client.on(events.NewMessage(pattern='(?i)/basla'))
async def handle_start(event): 
    await event.reply("🤖 Royal Baccarat Bot Aktif! 🎯")

@client.on(events.NewMessage(pattern='(?i)/durum'))
async def handle_durum(event):
    aktif_sinyal = "✅ Evet" if is_signal_active else "❌ Hayır"
    gmt3_time = datetime.now(GMT3).strftime('%H:%M:%S')
    
    total = performance_stats['total_signals']
    wins = performance_stats['win_signals']
    win_rate = (wins / total * 100) if total > 0 else 0
    
    await event.reply(f"""🤖 **ROYAL BACCARAT BOT** 🎯

🟢 **Durum:** Çalışıyor
🎯 **Aktif Sinyal:** {aktif_sinyal}
🎛️ **Mod:** {SISTEM_MODU}
🕒 **Saat:** {gmt3_time} (GMT+3)
📨 **Günlük Sinyal:** {daily_signal_count}

📊 **İSTATİSTİKLER:**
• Toplam Sinyal: {total}
• Kazanç: {wins} | Kayıp: {performance_stats['loss_signals']}
• Başarı Oranı: %{win_rate:.1f}
• Toplam Kâr: {performance_stats['total_profit']}
• Seri: {performance_stats['current_streak']} (Rekor: {performance_stats['max_streak']})""")

@client.on(events.NewMessage(pattern='(?i)/test'))
async def handle_test(event):
    try:
        debug_log("🧪 Test sinyali gönderiliyor...")
        await send_new_signal(999, '♠', "TEST SİNYALİ")
        await event.reply("✅ Test sinyali gönderildi!")
    except Exception as e:
        await event.reply(f"❌ Test hatası: {e}")

@client.on(events.NewMessage(pattern='(?i)/debug'))
async def handle_debug(event):
    active_signals = []
    for game_num, tracker in martingale_trackers.items():
        active_signals.append(f"#{game_num} - {tracker['signal_suit']} (Seviye: {tracker['step']})")
    
    debug_info = f"""🔧 **DEBUG BİLGİLERİ**

📊 Game Results: {len(game_results)} oyun
🎯 Active Signals: {len(martingale_trackers)}
🔔 Signal Active: {is_signal_active}
📨 Daily Signals: {daily_signal_count}

🔍 **Aktif Sinyaller:**
{chr(10).join(active_signals) if active_signals else 'Aktif sinyal yok'}"""
    
    await event.reply(debug_info)

# ANA MESAJ İŞLEYİCİ
@client.on(events.NewMessage(chats=KANAL_KAYNAK_ID))
async def handle_source_channel_message(event):
    try:
        text = event.message.text or ""
        debug_log(f"📥 Kaynak kanaldan mesaj: {text}")
        
        # Yeni format için regex
        game_match = re.search(r'№(\d+)', text)
        player_match = re.search(r'👉[^()]*\(([^)]+)\)', text)
        banker_match = re.search(r'-\s*[^()]*\(([^)]+)\)', text)
        
        debug_log(f"🔍 Parse sonuçları - Game: {game_match}, Player: {player_match}")
        
        if game_match and player_match:
            game_number = int(game_match.group(1))
            player_cards = player_match.group(1)
            banker_cards = banker_match.group(1) if banker_match else ''
            
            debug_log(f"🎮 Oyun #{game_number} - Player: {player_cards}")
            
            game_info = {
                'game_number': game_number,
                'player_cards': player_cards,
                'banker_cards': banker_cards,
                'is_final': True
            }
            
            # OYUNU KAYDET
            game_results[game_number] = game_info
            
            # Excel'e kaydet
            now = datetime.now(GMT3)
            game_data = {
                'game_number': game_number,
                'date': now.strftime('%d.%m.%Y'),
                'time': now.strftime('%H:%M:%S'),
                'player_cards': player_cards,
                'banker_cards': banker_cards,
                'winner': 'Tie',
                'color_prediction': 'N/A',
                'pattern_type': 'N/A',
                'result': 'N/A',
                'profit_loss': 0,
                'martingale_level': 0
            }
            save_to_excel(game_data)
            
            # SONUÇ KONTROLÜ
            await check_and_update_signal_results()
            
            # YENİ SİNYAL TETİKLEME
            if not is_signal_active:
                debug_log("🎯 Yeni sinyal tetikleniyor...")
                await normal_hibrit_sistemi(game_info)
                    
    except Exception as e: 
        debug_log(f"❌ Mesaj işleme hatası: {e}")

# ANA ÇALIŞTIRMA
async def main():
    try:
        debug_log("🚀 Bot başlatılıyor...")
        await client.start(bot_token=BOT_TOKEN)
        
        me = await client.get_me()
        debug_log(f"✅ Bot başarıyla başlatıldı: @{me.username}")
        
        # Kanal kontrolleri
        try:
            target_entity = await client.get_entity(KANAL_HEDEF)
            debug_log(f"✅ Hedef kanal: {target_entity.title}")
        except Exception as e:
            debug_log(f"❌ Hedef kanal hatası: {e}")

        try:
            source_entity = await client.get_entity(KANAL_KAYNAK_ID)
            debug_log(f"✅ Kaynak kanal: {source_entity.title}")
        except Exception as e:
            debug_log(f"❌ Kaynak kanal hatası: {e}")
        
        await client.run_until_disconnected()
        
    except Exception as e:
        debug_log(f"❌ Bot başlatma hatası: {e}")

if __name__ == '__main__':
    print("🤖 ROYAL BACCARAT BOT BAŞLIYOR...")
    print("🔍 DEBUG MOD: AKTİF")
    print("⚡ MARTINGALE SİSTEMİ: AKTİF")
    
    # Eski session dosyalarını temizle
    session_files = ['/root/0zctn/royal_bot_session.session', 'royal_bot.session']
    for session_file in session_files:
        if os.path.exists(session_file):
            try:
                os.remove(session_file)
                print(f"✅ Eski session silindi: {session_file}")
            except Exception as e:
                print(f"⚠️ Session silinemedi: {e}")
    
    init_excel_file()
    
    try:            
        client.loop.run_until_complete(main())
    except KeyboardInterrupt:
        print("🔴 Bot durduruldu")
    except Exception as e:
        print(f"❌ Bot hatası: {e}")
