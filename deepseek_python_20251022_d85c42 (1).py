# -*- coding: utf-8 -*-
import re, json, os, asyncio, sys, pytz, logging
from datetime import datetime, timedelta
from telethon import TelegramClient, events
from telethon.errors import FloodWaitError, MessageNotModifiedError, RPCError
from telethon.tl.types import DocumentAttributeFilename
from collections import defaultdict, deque
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

# Logging ayarı
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('/var/log/royal_bot.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger('RoyalBot')

# Config değişkenleri
API_ID = 29581698
API_HASH = '0caabd4263f1d4e5f753659a787c2e7d'
BOT_TOKEN = ''  # 🔑 Buraya bot tokenınızı yazın
KANAL_KAYNAK_ID = -1001626824569
KANAL_HEDEF = "@royalbaccfree"
ADMIN_ID = 1136442929
SISTEM_MODU = "normal_hibrit"
GMT3 = pytz.timezone('Europe/Istanbul')

# CLIENT TANIMI - EN BAŞTA
client = TelegramClient('royal_bot', API_ID, API_HASH)

# Global değişkenler
game_results, martingale_trackers, color_trend, recent_games = {}, {}, [], []
MAX_MARTINGALE_STEPS, MAX_GAME_NUMBER, is_signal_active, daily_signal_count = 3, 1440, False, 0

# Excel dosyası
EXCEL_FILE = "/root/0zctn/royal_baccarat_data.xlsx"
RED_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
BLUE_FILL = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
GREEN_FILL = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
BLACK_FILL = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

# C2_3 Tipleri
C2_3_TYPES = {
    '#C2_3': {'emoji': '🔴', 'name': 'KLASİK', 'confidence': 0.9, 'stats': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0}},
    '#C2_2': {'emoji': '🔵', 'name': 'ALTERNATİF', 'confidence': 0.7, 'stats': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0}},
    '#C3_2': {'emoji': '🟢', 'name': 'VARYANT', 'confidence': 0.6, 'stats': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0}},
    '#C3_3': {'emoji': '🟡', 'name': 'ÖZEL', 'confidence': 0.7, 'stats': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0}}
}

# İstatistikler
performance_stats = {
    'total_signals': 0, 'win_signals': 0, 'loss_signals': 0, 'total_profit': 0,
    'current_streak': 0, 'max_streak': 0,
    'daily_stats': defaultdict(lambda: {'signals': 0, 'wins': 0, 'losses': 0, 'profit': 0}),
    'weekly_stats': defaultdict(lambda: {'signals': 0, 'wins': 0, 'losses': 0, 'profit': 0}),
    'signal_history': deque(maxlen=1000),
    'c2_3_performance': C2_3_TYPES.copy()
}

pattern_stats = {
    '🎯 GÜÇLÜ EL': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0, 'avg_steps': 0},
    '🏆 DOĞAL KAZANÇ': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0, 'avg_steps': 0},
    '📊 5+ KART': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0, 'avg_steps': 0},
    '🚨 3x TEKRAR': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0, 'avg_steps': 0},
    '📈 STANDART SİNYAL': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0, 'avg_steps': 0},
    '✅ 5-Lİ ONAY': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0, 'avg_steps': 0},
    '🚀 SÜPER HİBRİT': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0, 'avg_steps': 0}
}

# EXCEL FONKSİYONLARI
def init_excel_file():
    try:
        if not os.path.exists(EXCEL_FILE):
            wb = Workbook()
            ws = wb.active
            ws.title = "Baccarat Data"
            headers = ["Oyun No", "Tarih", "Saat", "Player Kartları", "Banker Kartları", "Player Toplam", "Banker Toplam", "Kazanan", "C2 Tipi", "Renk Tahmini", "Pattern Tipi", "Sinyal Seviyesi", "Sonuç", "Kazanç/Kayıp", "Toplam Kâr"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header).font = Font(bold=True)
            wb.save(EXCEL_FILE)
            print("✅ Excel dosyası oluşturuldu")
    except Exception as e:
        print(f"❌ Excel oluşturma hatası: {e}")

def save_to_excel(game_data):
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        row = ws.max_row + 1
        
        ws.cell(row=row, column=1, value=game_data.get('game_number'))
        ws.cell(row=row, column=2, value=game_data.get('date'))
        ws.cell(row=row, column=3, value=game_data.get('time'))
        ws.cell(row=row, column=4, value=game_data.get('player_cards'))
        ws.cell(row=row, column=5, value=game_data.get('banker_cards'))
        ws.cell(row=row, column=6, value=game_data.get('player_total'))
        ws.cell(row=row, column=7, value=game_data.get('banker_total'))
        ws.cell(row=row, column=8, value=game_data.get('winner'))
        ws.cell(row=row, column=9, value=game_data.get('c2_type'))
        
        color_cell = ws.cell(row=row, column=10, value=game_data.get('color_prediction'))
        color_pred = game_data.get('color_prediction', '')
        if '🔴' in color_pred or 'MAÇA' in color_pred: color_cell.fill = RED_FILL
        elif '🔵' in color_pred or 'KALP' in color_pred: color_cell.fill = BLUE_FILL
        elif '🟢' in color_pred or 'KARO' in color_pred: color_cell.fill = GREEN_FILL
        elif '⚫' in color_pred or 'SİNEK' in color_pred: color_cell.fill = BLACK_FILL
        
        ws.cell(row=row, column=11, value=game_data.get('pattern_type'))
        ws.cell(row=row, column=12, value=game_data.get('signal_level'))
        
        result_cell = ws.cell(row=row, column=13, value=game_data.get('result'))
        if game_data.get('result') == 'KAZANÇ': result_cell.fill = GREEN_FILL
        elif game_data.get('result') == 'KAYIP': result_cell.fill = RED_FILL
            
        ws.cell(row=row, column=14, value=game_data.get('profit_loss'))
        ws.cell(row=row, column=15, value=game_data.get('total_profit'))
        
        wb.save(EXCEL_FILE)
    except Exception as e:
        print(f"❌ Excel kaydetme hatası: {e}")

async def save_game_result_to_excel(game_info):
    try:
        # Mevcut oyun numarasını al
        game_number = game_info['game_number']
        
        # Tarih ve saat bilgisi (GMT+3)
        now = datetime.now(GMT3)
        date_str = now.strftime('%d.%m.%Y')
        time_str = now.strftime('%H:%M:%S')
        
        # Kart değerlerini hesapla
        player_cards = game_info['player_cards']
        banker_cards = game_info['banker_cards']
        
        # Player toplamını hesapla
        player_total = 0
        player_card_matches = re.findall(r'(10|[A2-9TJQK])', player_cards)
        for card in player_card_matches:
            player_total += get_baccarat_value(card)
        player_total %= 10
        
        # Banker toplamını hesapla
        banker_total = 0
        banker_card_matches = re.findall(r'(10|[A2-9TJQK])', banker_cards)
        for card in banker_card_matches:
            banker_total += get_baccarat_value(card)
        banker_total %= 10
        
        # Kazananı belirle
        if player_total > banker_total:
            winner = "Player"
        elif banker_total > player_total:
            winner = "Banker"
        else:
            winner = "Tie"
        
        # Renk tahmini (en yüksek değerli kartın rengi)
        signal_color = extract_largest_value_suit(player_cards)
        color_display = get_suit_display_name(signal_color) if signal_color else "N/A"
        
        # Pattern analizi
        pattern_color, pattern_type = analyze_simple_pattern(player_cards, banker_cards, game_number)
        
        # Excel'e kaydetmek için veri hazırla
        game_data = {
            'game_number': game_number,
            'date': date_str,
            'time': time_str,
            'player_cards': player_cards,
            'banker_cards': banker_cards,
            'player_total': player_total,
            'banker_total': banker_total,
            'winner': winner,
            'c2_type': game_info.get('c2_3_type', 'N/A'),
            'color_prediction': color_display,
            'pattern_type': pattern_type,
            'signal_level': 'N/A',
            'result': 'N/A',
            'profit_loss': 0,
            'total_profit': 0
        }
        
        # Excel'e kaydet
        save_to_excel(game_data)
        logger.info(f"✅ Oyun #{game_number} Excel'e kaydedildi")
        
    except Exception as e:
        logger.error(f"❌ Oyun kaydetme hatası: {e}")

# TEMEL FONKSİYONLAR
def get_suit_display_name(suit_symbol):
    suit_names = {'♠': '♠️ MAÇA', '♥': '❤️ KALP', '♦': '♦️ KARO', '♣': '♣️ SİNEK'}
    return suit_names.get(suit_symbol, f"❓ {suit_symbol}")

def get_baccarat_value(card_char):
    if card_char == '10': return 10
    if card_char in 'AKQJ2T': return 0
    elif card_char.isdigit(): return int(card_char)
    return -1

def extract_largest_value_suit(cards_str):
    try:
        cards = re.findall(r'(10|[A2-9TJQK])([♣♦♥♠])', cards_str)
        if not cards: return None
        max_value, largest_value_suit = -1, None
        values = [get_baccarat_value(card[0]) for card in cards]
        if len(values) == 2 and values[0] == values[1]: return None
        for card_char, suit in cards:
            value = get_baccarat_value(card_char)
            if value > max_value: max_value, largest_value_suit = value, suit
        return largest_value_suit if max_value > 0 else None
    except Exception as e:
        print(f"❌ extract_largest_value_suit hatası: {e}")
        return None

def analyze_simple_pattern(player_cards, banker_cards, game_number):
    try:
        signal_color = extract_largest_value_suit(player_cards)
        if not signal_color: return None, "Renk tespit edilemedi"
        color_trend.append(signal_color)
        if len(color_trend) > 10: color_trend.pop(0)
        player_card_data = re.findall(r'(10|[A2-9TJQK])([♣♦♥♠])', player_cards)
        player_values = [get_baccarat_value(card[0]) for card in player_card_data]
        banker_card_data = re.findall(r'(10|[A2-9TJQK])([♣♦♥♠])', banker_cards)
        banker_values = [get_baccarat_value(card[0]) for card in banker_card_data]
        total_cards = len(player_values) + len(banker_values)
        if sum(player_values) >= 8 and len(player_values) >= 3: return signal_color, "🎯 GÜÇLÜ EL"
        elif sum(player_values) in [8, 9]: return signal_color, "🏆 DOĞAL KAZANÇ"
        elif total_cards >= 5: return signal_color, "📊 5+ KART"
        elif len(color_trend) >= 3 and color_trend[-3:] == [signal_color] * 3: return signal_color, "🚨 3x TEKRAR"
        else: return signal_color, "📈 STANDART SİNYAL"
    except Exception as e:
        return None, f"Hata: {e}"

def get_next_game_number(current_game_num):
    next_num = current_game_num + 1
    return 1 if next_num > MAX_GAME_NUMBER else next_num

# EXCEL ANALİZ FONKSİYONLARI
def analyze_cards_by_c2_type():
    try:
        if not os.path.exists(EXCEL_FILE): return "❌ Excel dosyası bulunamadı"
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        if ws.max_row <= 1: return "📊 Excel'de henüz veri yok"

        c2_card_stats, c2_suit_stats, c2_value_stats = {}, {}, {}
        for row in range(2, ws.max_row + 1):
            c2_type = ws.cell(row=row, column=9).value
            player_cards = ws.cell(row=row, column=4).value
            banker_cards = ws.cell(row=row, column=5).value
            result = ws.cell(row=row, column=13).value

            if c2_type and c2_type != 'N/A' and player_cards and banker_cards:
                if c2_type not in c2_card_stats:
                    c2_card_stats[c2_type] = {'player_cards': [], 'banker_cards': [], 'total_games': 0, 'wins': 0, 'losses': 0}
                    c2_suit_stats[c2_type] = {'♠': 0, '♥': 0, '♦': 0, '♣': 0}
                    c2_value_stats[c2_type] = defaultdict(int)

                c2_card_stats[c2_type]['player_cards'].append(player_cards)
                c2_card_stats[c2_type]['banker_cards'].append(banker_cards)
                c2_card_stats[c2_type]['total_games'] += 1
                if result == 'KAZANÇ': c2_card_stats[c2_type]['wins'] += 1
                elif result == 'KAYIP': c2_card_stats[c2_type]['losses'] += 1

                all_cards_text = player_cards + banker_cards
                suits = re.findall(r'[♠♥♦♣]', all_cards_text)
                for suit in suits: c2_suit_stats[c2_type][suit] += 1

                card_values = re.findall(r'(10|[A2-9TJQK])', all_cards_text)
                for value in card_values: c2_value_stats[c2_type][value] += 1

        return generate_card_analysis_report(c2_card_stats, c2_suit_stats, c2_value_stats)
    except Exception as e: return f"❌ Kart analiz hatası: {e}"

def generate_card_analysis_report(c2_card_stats, c2_suit_stats, c2_value_stats):
    if not c2_card_stats: return "📊 C2 tipine göre kart verisi bulunamadı"
    report = "🃏 **C2 TİPİNE GÖRE KART ANALİZİ** 🃏\n\n"
    for c2_type, stats in c2_card_stats.items():
        total_games = stats['total_games']
        win_rate = (stats['wins'] / total_games * 100) if total_games > 0 else 0
        report += f"🎯 **{c2_type}** - {total_games} oyun (%{win_rate:.1f} başarı)\n"
        suit_stats = c2_suit_stats[c2_type]
        total_suits = sum(suit_stats.values())
        report += "🎨 **Renk Dağılımı:**\n"
        for suit, count in suit_stats.items():
            if total_suits > 0:
                percentage = (count / total_suits) * 100
                suit_name = get_suit_display_name(suit)
                report += f"   {suit_name}: {count} (%{percentage:.1f})\n"
        value_stats = c2_value_stats[c2_type]
        total_values = sum(value_stats.values())
        report += "🔢 **Kart Değerleri (Top 5):**\n"
        sorted_values = sorted(value_stats.items(), key=lambda x: x[1], reverse=True)[:5]
        for value, count in sorted_values:
            if total_values > 0:
                percentage = (count / total_values) * 100
                report += f"   {value}: {count} (%{percentage:.1f})\n"
        report += "\n" + "─" * 40 + "\n\n"
    return report

def get_c2_recommendation_from_excel():
    try:
        if not os.path.exists(EXCEL_FILE): return "#C2_3", "Excel dosyası bulunamadı"
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        if ws.max_row <= 10: return "#C2_3", "Yeterli veri yok"
        recent_games = min(30, ws.max_row - 1)
        c2_recent_stats = {}
        for row in range(ws.max_row - recent_games + 1, ws.max_row + 1):
            c2_type = ws.cell(row=row, column=9).value
            result = ws.cell(row=row, column=13).value
            if c2_type and c2_type != 'N/A':
                if c2_type not in c2_recent_stats: c2_recent_stats[c2_type] = {'total': 0, 'wins': 0}
                c2_recent_stats[c2_type]['total'] += 1
                if result == 'KAZANÇ': c2_recent_stats[c2_type]['wins'] += 1
        if not c2_recent_stats: return "#C2_3", "C2 verisi yok"
        best_c2 = max(c2_recent_stats.items(), key=lambda x: (x[1]['wins']/x[1]['total']) if x[1]['total'] > 0 else 0)
        c2_type, stats = best_c2
        win_rate = (stats['wins'] / stats['total']) * 100 if stats['total'] > 0 else 0
        return c2_type, f"Son {recent_games} oyunda %{win_rate:.1f} başarı"
    except Exception as e: return "#C2_3", f"Hata: {e}"

def predict_color_by_c2_excel(c2_type, player_cards):
    try:
        if not os.path.exists(EXCEL_FILE): return predict_color_by_c2_type(c2_type, player_cards)
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        if ws.max_row <= 5: return predict_color_by_c2_type(c2_type, player_cards)
        c2_color_stats = {}
        for row in range(2, ws.max_row + 1):
            row_c2_type = ws.cell(row=row, column=9).value
            color_pred = ws.cell(row=row, column=10).value
            result = ws.cell(row=row, column=13).value
            if row_c2_type == c2_type and color_pred and result in ['KAZANÇ', 'KAYIP']:
                for color in ['🔴', '🔵', '🟢', '⚫']:
                    if color in str(color_pred):
                        if color not in c2_color_stats: c2_color_stats[color] = {'total': 0, 'wins': 0}
                        c2_color_stats[color]['total'] += 1
                        if result == 'KAZANÇ': c2_color_stats[color]['wins'] += 1
                        break
        if c2_color_stats:
            best_color = max(c2_color_stats.items(), key=lambda x: (x[1]['wins']/x[1]['total']) if x[1]['total'] > 0 else 0)
            color_emoji, color_stats = best_color
            win_rate = (color_stats['wins'] / color_stats['total']) * 100 if color_stats['total'] > 0 else 0
            base_color = extract_largest_value_suit(player_cards)
            suit_display = get_suit_display_name(base_color) if base_color else "MAÇA"
            return f"{color_emoji} {suit_display}", f"Excel C2 Optimize: %{win_rate:.1f} başarı"
        return predict_color_by_c2_type(c2_type, player_cards)
    except Exception as e:
        return predict_color_by_c2_type(c2_type, player_cards)

def predict_color_by_c2_type(c2_type, player_cards):
    try:
        base_color = extract_largest_value_suit(player_cards)
        if not base_color: return "🔴 MAÇA", "Varsayılan"
        suit_display = get_suit_display_name(base_color)
        if c2_type == '#C2_3': return f"🔴 {suit_display}", "KLASİK C2_3 - KIRMIZI agresif"
        elif c2_type == '#C2_2': return f"🔵 {suit_display}", "ALTERNATİF C2_2 - MAVİ dengeli"
        elif c2_type == '#C3_2': return f"🟢 {suit_display}", "VARYANT C3_2 - YEŞIL riskli"
        elif c2_type == '#C3_3': return f"🟡 {suit_display}", "ÖZEL C3_3 - SARI özel"
        else: return f"🔴 {suit_display}", "Varsayılan tahmin"
    except Exception as e: return "🔴 MAÇA", "Hata"

# SİNYAL SİSTEMLERİ
async def send_new_signal(game_num, signal_suit, reason, c2_3_info=None):
    global is_signal_active, daily_signal_count
    try:
        suit_display = get_suit_display_name(signal_suit)
        if c2_3_info:
            c2_3_type, c2_3_desc = c2_3_info.get('c2_3_type', ''), c2_3_info.get('c2_3_description', '')
            trigger_info = f"{c2_3_desc} {c2_3_type}"
        else: 
            c2_3_type, c2_3_desc = '#C2_3', 'KLASİK'
            trigger_info = "KLASİK #C2_3"
        
        gmt3_time = datetime.now(GMT3).strftime('%H:%M:%S')
        signal_text = f"🎯 **SİNYAL BAŞLADI** 🎯\n#N{game_num} - {suit_display}\n📊 Tetikleyici: {trigger_info}\n🎯 Sebep: {reason}\n⚡ Strateji: Martingale {MAX_MARTINGALE_STEPS} Seviye\n🕒 {gmt3_time} (GMT+3)\n🔴 SONUÇ: BEKLENİYOR..."
        sent_message = await client.send_message(KANAL_HEDEF, signal_text)
        daily_signal_count += 1
        martingale_trackers[game_num] = {
            'message_obj': sent_message, 'step': 0, 'signal_suit': signal_suit, 
            'sent_game_number': game_num, 'expected_game_number_for_check': game_num, 
            'start_time': datetime.now(GMT3), 'reason': reason, 
            'c2_3_type': c2_3_type, 'c2_3_description': c2_3_desc, 'results': []
        }
        is_signal_active = True
    except Exception as e: print(f"❌ Sinyal gönderme hatası: {e}")

async def normal_hibrit_sistemi(game_info):
    trigger_game_num, c2_3_info = game_info['game_number'], {'c2_3_type': game_info.get('c2_3_type'), 'c2_3_description': game_info.get('c2_3_description')}
    signal_color, reason = analyze_simple_pattern(game_info['player_cards'], game_info['banker_cards'], trigger_game_num)
    if signal_color:
        next_game_num = get_next_game_number(trigger_game_num)
        await send_new_signal(next_game_num, signal_color, reason, c2_3_info)

async def excel_based_sistemi(game_info):
    try:
        recommended_c2, c2_reason = get_c2_recommendation_from_excel()
        smart_color, smart_reason = predict_color_by_c2_excel(recommended_c2, game_info['player_cards'])
        next_game_num = get_next_game_number(game_info['game_number'])
        c2_3_info = {'c2_3_type': recommended_c2, 'c2_3_description': f"EXCEL TAVSİYE - {c2_reason}"}
        await send_new_signal(next_game_num, smart_color[0], f"📊 EXCEL SİSTEM - {smart_reason}", c2_3_info)
    except Exception as e: print(f"❌ Excel sistemi hatası: {e}")

# KOMUTLAR
@client.on(events.NewMessage(pattern='(?i)/basla'))
async def handle_start(event): 
    await event.reply("🤖 Royal Baccarat Bot Aktif! 🎯")

@client.on(events.NewMessage(pattern='(?i)/durum'))
async def handle_durum(event):
    aktif_sinyal = "✅ Evet" if is_signal_active else "❌ Hayır"
    gmt3_time = datetime.now(GMT3).strftime('%H:%M:%S')
    await event.reply(f"""🤖 **ROYAL BACCARAT BOT** 🎯

🟢 **Durum:** Çalışıyor
🎯 **Aktif Sinyal:** {aktif_sinyal}
🎛️ **Mod:** {SISTEM_MODU}
🕒 **Saat:** {gmt3_time} (GMT+3)
📨 **Günlük Sinyal:** {daily_signal_count}""")

@client.on(events.NewMessage(pattern='(?i)/excel_gonder'))
async def handle_excel_gonder(event):
    try:
        if event.sender_id != ADMIN_ID: 
            await event.reply("❌ Bu komut sadece admin için!")
            return
        if not os.path.exists(EXCEL_FILE):
            await event.reply("❌ Excel dosyası bulunamadı!")
            return
        file_size = os.path.getsize(EXCEL_FILE)
        file_size_mb = file_size / (1024 * 1024)
        if file_size_mb > 50:
            await event.reply(f"❌ Excel dosyası çok büyük ({file_size_mb:.1f}MB)")
            return
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        total_records = ws.max_row - 1
        await event.reply(f"📊 **Excel Gönderiliyor...**\n📈 Kayıt: {total_records}\n💾 Boyut: {file_size_mb:.1f}MB")
        await client.send_file(event.chat_id, EXCEL_FILE, caption=f"📈 **Royal Baccarat Veri Tablosu**\n📊 {total_records} kayıt\n🕒 {datetime.now(GMT3).strftime('%d.%m.%Y %H:%M')}")
    except Exception as e: await event.reply(f"❌ Excel gönderme hatası: {e}")

@client.on(events.NewMessage(pattern='(?i)/kart_analiz'))
async def handle_kart_analiz(event):
    analysis = analyze_cards_by_c2_type()
    await event.reply(analysis)

@client.on(events.NewMessage(pattern='(?i)/excel_tahmin'))
async def handle_excel_tahmin(event):
    try:
        recommended_c2, c2_reason = get_c2_recommendation_from_excel()
        smart_color, smart_reason = predict_color_by_c2_excel(recommended_c2, "")
        await event.reply(f"""🎯 **EXCEL TABANLI TAHMİN** 🎯

📊 **Excel Analizi:**
• Önerilen C2 Tipi: {recommended_c2}
• Sebep: {c2_reason}

🎨 **TAHMİN:**
• Renk: {smart_color}
• Sebep: {smart_reason}

💡 **STRATEJİ:** {smart_color.split()[0]} odaklanın!""")
    except Exception as e: await event.reply(f"❌ Excel tahmin hatası: {e}")

@client.on(events.NewMessage(pattern='(?i)/mod_excel'))
async def handle_mod_excel(event):
    global SISTEM_MODU
    SISTEM_MODU = "excel_based"
    record_count = 0
    if os.path.exists(EXCEL_FILE):
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active
            record_count = ws.max_row - 1
        except: pass
    await event.reply(f"""📊 **EXCEL TABANLI MOD AKTİF!**

📈 Kayıtlı Veri: {record_count} oyun

🎯 **ÖZELLİKLER:**
• Excel'deki geçmiş verileri analiz eder
• En başarılı C2 tipini otomatik seçer
• Gerçek istatistiklere dayalı tahmin yapar""")

@client.on(events.NewMessage(pattern='(?i)/mod_normal'))
async def handle_mod_normal(event):
    global SISTEM_MODU
    SISTEM_MODU = "normal_hibrit"
    await event.reply("✅ **NORMAL HİBRİT MOD AKTİF!**")

@client.on(events.NewMessage(pattern='(?i)/yardim'))
async def handle_yardim(event):
    yardim = """🤖 **ROYAL BACCARAT BOT** 🎯

**TEMEL KOMUTLAR:**
• /basla - Botu başlat
• /durum - Sistem durumu

**EXCEL KOMUTLARI:**
• /excel_gonder - 📁 Excel dosyasını gönder (ADMIN)
• /kart_analiz - C2 tipi kart analizi
• /excel_tahmin - Excel tabanlı tahmin

**SİSTEM MODLARI:**
• /mod_normal - Normal Hibrit Mod
• /mod_excel - 📊 Excel Tabanlı Mod

💾 **Excel Özellikleri:**
• Tüm oyunlar otomatik kaydedilir
• C2 tipi analizleri
• Gerçek verilere dayalı tahminler"""
    await event.reply(yardim)

# ANA MESAJ İŞLEYİCİ
@client.on(events.NewMessage(chats=KANAL_KAYNAK_ID))
async def handle_source_channel_message(event):
    try:
        text = event.message.text or ""
        game_match = re.search(r'#N(\d+)', text)
        player_match = re.search(r'\((.*?)\)', text)
        banker_match = re.search(r'\d+\s+\((.*?)\)', text)
        
        if game_match and player_match:
            game_info = {
                'game_number': int(game_match.group(1)),
                'player_cards': player_match.group(1),
                'banker_cards': banker_match.group(1) if banker_match else '',
                'is_final': True,
                'is_c2_3': any(trigger in text for trigger in C2_3_TYPES.keys()),
                'c2_3_type': next((trigger for trigger in C2_3_TYPES.keys() if trigger in text), '#C2_3'),
                'c2_3_description': C2_3_TYPES.get(next((trigger for trigger in C2_3_TYPES.keys() if trigger in text), '#C2_3'), {}).get('name', 'KLASİK')
            }
            
            game_results[game_info['game_number']] = game_info
            await save_game_result_to_excel(game_info)
            
            if not is_signal_active and game_info['is_final'] and game_info.get('is_c2_3'):
                if SISTEM_MODU == "normal_hibrit": 
                    await normal_hibrit_sistemi(game_info)
                elif SISTEM_MODU == "excel_based":
                    await excel_based_sistemi(game_info)
    except Exception as e: print(f"❌ Mesaj işleme hatası: {e}")

# ANA ÇALIŞTIRMA
if __name__ == '__main__':
    print("🤖 ROYAL BACCARAT BOT BAŞLIYOR...")
    init_excel_file()
    print("⏳ Bağlanıyor...")
    try:
        with client: 
            client.start(bot_token=BOT_TOKEN)
            client.run_until_disconnected()
    except Exception as e: 
        print(f"❌ Bot hatası: {e}")