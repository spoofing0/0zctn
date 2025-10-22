# -*- coding: utf-8 -*-
import re, json, os, asyncio, sys, pytz
from datetime import datetime, timedelta
from telethon import TelegramClient, events
from telethon.errors import FloodWaitError, MessageNotModifiedError
from collections import defaultdict, deque
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import random

API_ID = 29581698
API_HASH = '0caabd4263f1d4e5f753659a787c2e7d'
BOT_TOKEN = ''  # 🔑 Buraya bot tokenınızı yazın
KANAL_KAYNAK_ID = -1001626824569
KANAL_HEDEF = "@royalbaccfree"  # 📢 Hedef kanal
ADMIN_ID = 1136442929  # 👑 Admin ID
SISTEM_MODU = "normal_hibrit"
GMT3 = pytz.timezone('Europe/Istanbul')
client = TelegramClient('royal_bot', API_ID, API_HASH).start(bot_token=BOT_TOKEN)

game_results, martingale_trackers, color_trend, recent_games = {}, {}, [], []
MAX_MARTINGALE_STEPS, MAX_GAME_NUMBER, is_signal_active, daily_signal_count = 3, 1440, False, 0

# Excel dosyası için değişkenler
EXCEL_FILE = "baccarat_data.xlsx"
excel_data = []

# Güncellenmiş C2_3 istatistik yapısı
C2_3_TYPES = {
    '#C2_3': {'emoji': '🔴', 'name': 'KLASİK', 'confidence': 0.9, 'stats': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0}},
    '#C2_2': {'emoji': '🔵', 'name': 'ALTERNATİF', 'confidence': 0.7, 'stats': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0}},
    '#C3_2': {'emoji': '🟢', 'name': 'VARYANT', 'confidence': 0.6, 'stats': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0}},
    '#C3_3': {'emoji': '🟡', 'name': 'ÖZEL', 'confidence': 0.7, 'stats': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0}}
}

# İstatistik veri yapıları
performance_stats = {
    'total_signals': 0,
    'win_signals': 0,
    'loss_signals': 0,
    'total_profit': 0,
    'current_streak': 0,
    'max_streak': 0,
    'daily_stats': defaultdict(lambda: {'signals': 0, 'wins': 0, 'losses': 0, 'profit': 0}),
    'weekly_stats': defaultdict(lambda: {'signals': 0, 'wins': 0, 'losses': 0, 'profit': 0}),
    'signal_history': deque(maxlen=1000),
    'c2_3_performance': C2_3_TYPES.copy(),
    'excel_stats': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0}
}

# PATTERN İSTATİSTİKLERİ
pattern_stats = {
    '🎯 GÜÇLÜ EL': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0, 'avg_steps': 0},
    '🏆 DOĞAL KAZANÇ': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0, 'avg_steps': 0},
    '📊 5+ KART': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0, 'avg_steps': 0},
    '🚨 3x TEKRAR': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0, 'avg_steps': 0},
    '📈 STANDART SİNYAL': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0, 'avg_steps': 0},
    '✅ 5-Lİ ONAY': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0, 'avg_steps': 0},
    '🚀 SÜPER HİBRİT': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0, 'avg_steps': 0},
    '🎯 KLASİK #C2_3': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0, 'avg_steps': 0},
    '📊 EXCEL TAHMİN': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0, 'avg_steps': 0}
}

# TEMEL FONKSİYONLAR - EKSİK OLANLARI EKLEYELİM
def get_suit_display_name(suit_symbol):
    suit_names = {'♠': '♠️ MAÇA', '♥': '❤️ KALP', '♦': '♦️ KARO', '♣': '♣️ SİNEK'}
    return suit_names.get(suit_symbol, f"❓ {suit_symbol}")

def get_baccarat_value(card_char):
    if card_char == '10': return 10
    if card_char in 'AKQJ2T': return 0
    elif card_char.isdigit(): return int(card_char)
    return -1

def extract_largest_value_suit(cards_str):
    try:
        cards = re.findall(r'(10|[A2-9TJQK])([♣♦♥♠])', cards_str)
        if not cards: return None
        max_value, largest_value_suit = -1, None
        values = [get_baccarat_value(card[0]) for card in cards]
        if len(values) == 2 and values[0] == values[1]: return None
        for card_char, suit in cards:
            value = get_baccarat_value(card_char)
            if value > max_value: max_value, largest_value_suit = value, suit
        return largest_value_suit if max_value > 0 else None
    except Exception as e:
        print(f"❌ extract_largest_value_suit hatası: {e}")
        return None

def analyze_simple_pattern(player_cards, banker_cards, game_number):
    try:
        signal_color = extract_largest_value_suit(player_cards)
        if not signal_color: return None, "Renk tespit edilemedi"
        color_trend.append(signal_color)
        if len(color_trend) > 10: color_trend.pop(0)
        player_card_data = re.findall(r'(10|[A2-9TJQK])([♣♦♥♠])', player_cards)
        player_values = [get_baccarat_value(card[0]) for card in player_card_data]
        banker_card_data = re.findall(r'(10|[A2-9TJQK])([♣♦♥♠])', banker_cards)
        banker_values = [get_baccarat_value(card[0]) for card in banker_card_data]
        total_cards = len(player_values) + len(banker_values)
        if sum(player_values) >= 8 and len(player_values) >= 3: return signal_color, "🎯 GÜÇLÜ EL"
        elif sum(player_values) in [8, 9]: return signal_color, "🏆 DOĞAL KAZANÇ"
        elif total_cards >= 5: return signal_color, "📊 5+ KART"
        elif len(color_trend) >= 3 and color_trend[-3:] == [signal_color] * 3: return signal_color, "🚨 3x TEKRAR"
        else: return signal_color, "📈 STANDART SİNYAL"
    except Exception as e:
        print(f"❌ Pattern analysis error: {e}")
        return None, f"Hata: {e}"

def besli_onay_sistemi(player_cards, banker_cards, game_number):
    onaylar = []
    temel_renk = extract_largest_value_suit(player_cards)
    if temel_renk: 
        onaylar.append(("C2_3", temel_renk))
        print(f"✅ C2_3 onay: {temel_renk}")
    pattern_renk, pattern_sebep = analyze_simple_pattern(player_cards, banker_cards, game_number)
    if pattern_renk and "STANDART" not in pattern_sebep:
        onaylar.append(("PATTERN", pattern_renk))
        print(f"✅ Pattern onay: {pattern_renk} - {pattern_sebep}")
    if color_trend:
        trend_renk = color_trend[-1] if color_trend else None
        if trend_renk: onaylar.append(("TREND", trend_renk))
    if len(color_trend) >= 3:
        son_uc = color_trend[-3:]
        if son_uc.count(temel_renk) >= 2: onaylar.append(("REPEAT", temel_renk))
    renk_oyları = {}
    for yontem, renk in onaylar: renk_oyları[renk] = renk_oyları.get(renk, 0) + 1
    if renk_oyları:
        kazanan_renk = max(renk_oyları, key=renk_oyları.get)
        oy_sayisi = renk_oyları[kazanan_renk]
        güven = oy_sayisi / 5
        print(f"📊 5'li onay: {kazanan_renk} - {oy_sayisi}/5 - %{güven*100:.1f}")
        if oy_sayisi >= 3 and güven >= 0.6: return kazanan_renk, f"✅ 5-Lİ ONAY ({oy_sayisi}/5) - %{güven*100:.1f}"
    return None, "❌ 5'li onay sağlanamadı"

def super_filtre_kontrol(signal_color, reason, game_number):
    if len(color_trend) >= 5:
        if color_trend[-5:].count(signal_color) == 0: return False, "❌ SOĞUK TREND"
    if len(recent_games) >= 3:
        son_kayiplar = sum(1 for oyun in recent_games[-3:] if not oyun.get('kazanç', True))
        if son_kayiplar >= 2: return False, "🎯 ARDIŞIK KAYIPLAR"
    return True, "✅ TÜM FİLTRELER GEÇTİ"

def super_risk_analizi():
    risk_puan, uyarılar = 0, []
    if len(color_trend) >= 5:
        son_5 = color_trend[-5:]
        if len(set(son_5)) == 1: risk_puan, uyarılar = risk_puan + 30, uyarılar + ["🚨 5x AYNI RENK"]
    if risk_puan >= 30: return "🔴 YÜKSEK RİSK", uyarılar
    elif risk_puan >= 20: return "🟡 ORTA RİSK", uyarılar
    else: return "🟢 DÜŞÜK RİSK", uyarılar

def get_next_game_number(current_game_num):
    next_num = current_game_num + 1
    return 1 if next_num > MAX_GAME_NUMBER else next_num

# EKSİK FONKSİYONU EKLEYELİM
def extract_game_info_from_message(text):
    game_info = {'game_number': None, 'player_cards': '', 'banker_cards': '', 'is_final': False, 'is_player_drawing': False, 'is_c2_3': False, 'c2_3_type': None, 'c2_3_description': ''}
    try:
        game_match = re.search(r'#N(\d+)', text)
        if game_match: game_info['game_number'] = int(game_match.group(1))
        player_match = re.search(r'\((.*?)\)', text)
        if player_match: game_info['player_cards'] = player_match.group(1)
        banker_match = re.search(r'\d+\s+\((.*?)\)', text)
        if banker_match: game_info['banker_cards'] = banker_match.group(1)
        for trigger_type, trigger_data in C2_3_TYPES.items():
            if trigger_type in text:
                game_info['is_c2_3'], game_info['c2_3_type'], game_info['c2_3_description'] = True, trigger_type, trigger_data['name']
                break
        if ('✅' in text or '🔰' in text or '#X' in text): game_info['is_final'] = True
    except Exception as e: print(f"❌ Oyun bilgisi çıkarma hatası: {e}")
    return game_info

# EXCEL FONKSİYONLARI
def init_excel_file():
    """Excel dosyasını başlat"""
    try:
        if not os.path.exists(EXCEL_FILE):
            wb = Workbook()
            ws = wb.active
            ws.title = "Baccarat Data"
            # Başlıklar
            headers = [
                "Tarih", "Saat", "Oyun No", "Oyuncu Kartları", "Banker Kartları", 
                "Oyuncu Toplam", "Banker Toplam", "Kazanan", "Sinyal Rengi",
                "Tahmin Rengi", "Sonuç", "Martingale Seviye", "Kazanç/Kayıp"
            ]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            wb.save(EXCEL_FILE)
            print("✅ Excel dosyası oluşturuldu")
        else:
            print("✅ Excel dosyası zaten mevcut")
    except Exception as e:
        print(f"❌ Excel dosyası oluşturma hatası: {e}")

def calculate_hand_total(cards_str):
    """Kart toplamını hesapla"""
    try:
        cards = re.findall(r'(10|[A2-9TJQK])([♣♦♥♠])', cards_str)
        total = 0
        for card_char, suit in cards:
            value = get_baccarat_value(card_char)
            total += value
        return total % 10
    except:
        return 0

def determine_winner(player_total, banker_total):
    """Kazananı belirle"""
    if player_total > banker_total:
        return "PLAYER"
    elif banker_total > player_total:
        return "BANKER"
    else:
        return "TIE"

def save_to_excel(game_info, signal_color=None, result=None, martingale_step=0, profit=0):
    """Excel'e veri kaydet"""
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        # Son satırı bul
        row = ws.max_row + 1
        
        # Tarih ve saat
        now = datetime.now(GMT3)
        date_str = now.strftime('%Y-%m-%d')
        time_str = now.strftime('%H:%M:%S')
        
        # Oyuncu ve banker toplamını hesapla
        player_total = calculate_hand_total(game_info['player_cards'])
        banker_total = calculate_hand_total(game_info['banker_cards'])
        
        # Kazananı belirle
        winner = determine_winner(player_total, banker_total)
        
        # Tahmin rengi (eğer sinyal varsa)
        predicted_color = signal_color if signal_color else "YOK"
        
        # Sonuç
        result_text = result if result else "BEKLENİYOR"
        
        # Verileri yaz
        data = [
            date_str, time_str, game_info['game_number'], 
            game_info['player_cards'], game_info['banker_cards'],
            player_total, banker_total, winner, 
            extract_largest_value_suit(game_info['player_cards']),
            predicted_color, result_text, martingale_step, profit
        ]
        
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
        
        wb.save(EXCEL_FILE)
        print(f"✅ Excel'e kaydedildi: #{game_info['game_number']}")
        
    except Exception as e:
        print(f"❌ Excel kayıt hatası: {e}")

def analyze_excel_pattern():
    """Excel verilerine göre tahmin yap"""
    try:
        if not os.path.exists(EXCEL_FILE):
            return None, "Excel dosyası bulunamadı"
        
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        if ws.max_row <= 1:
            return None, "Yeterli veri yok"
        
        # Son 50 kaydı analiz et
        data = []
        for row in range(max(2, ws.max_row - 49), ws.max_row + 1):
            signal_color = ws.cell(row=row, column=9).value  # Sinyal Rengi sütunu
            result = ws.cell(row=row, column=11).value  # Sonuç sütunu
            
            if signal_color and result in ['KAZANÇ', 'KAYIP']:
                data.append({
                    'color': signal_color,
                    'win': result == 'KAZANÇ'
                })
        
        if len(data) < 10:
            return None, f"Yeterli veri yok ({len(data)} kayıt)"
        
        # Renk analizi
        color_stats = {}
        for suit in ['♠', '♥', '♦', '♣']:
            color_data = [d for d in data if d['color'] == suit]
            if color_data:
                wins = sum(1 for d in color_data if d['win'])
                total = len(color_data)
                win_rate = wins / total if total > 0 else 0
                color_stats[suit] = {
                    'total': total,
                    'wins': wins,
                    'win_rate': win_rate,
                    'recent_count': sum(1 for d in data[-10:] if d['color'] == suit)
                }
        
        if not color_stats:
            return None, "Renk verisi bulunamadı"
        
        # En iyi rengi seç (kazanç oranı + frekans)
        best_color = None
        best_score = -1
        
        for suit, stats in color_stats.items():
            # Skor: kazanç oranı * 0.7 + son 10'daki frekans * 0.3
            score = (stats['win_rate'] * 0.7) + ((stats['recent_count'] / 10) * 0.3)
            
            if score > best_score and stats['total'] >= 3:
                best_score = score
                best_color = suit
        
        if best_color and best_score >= 0.4:
            win_rate = color_stats[best_color]['win_rate'] * 100
            return best_color, f"📊 EXCEL TAHMİN (%{win_rate:.1f} başarı)"
        
        return None, f"Yeterli güven yok (en iyi skor: {best_score:.2f})"
        
    except Exception as e:
        return None, f"Analiz hatası: {e}"

def get_excel_performance():
    """Excel tahmin performansını getir"""
    stats = performance_stats['excel_stats']
    if stats['total'] == 0:
        return "📊 Excel Tahmin: Henüz veri yok"
    
    win_rate = (stats['wins'] / stats['total']) * 100
    return f"""📊 **EXCEL TAHMİN PERFORMANSI**

• Toplam Sinyal: {stats['total']}
• Kazanç: {stats['wins']} | Kayıp: {stats['losses']}
• Başarı Oranı: %{win_rate:.1f}
• Toplam Kâr: {stats['profit']} birim
"""

# EXCEL TAHMİN SİSTEMİ
async def excel_tahmin_sistemi(game_info):
    """Excel tabanlı tahmin sistemi"""
    print("📊 EXCEL TAHMİN analiz başlıyor...")
    
    # Excel'e mevcut oyunu kaydet (sonuçsuz)
    save_to_excel(game_info)
    
    # Excel analizi yap
    tahmin_renk, tahmin_sebep = analyze_excel_pattern()
    
    if not tahmin_renk:
        print(f"🚫 Excel Tahmin: {tahmin_sebep}")
        return
    
    # Risk kontrolü
    risk_seviye, risk_uyarilar = super_risk_analizi()
    if risk_seviye == "🔴 YÜKSEK RİSK":
        print(f"🚫 Excel Tahmin: Yüksek risk - {risk_uyarilar}")
        return
    
    # Günlük performans kontrolü
    daily = get_daily_stats()
    if daily['profit'] <= -10:
        print("🚫 Excel Tahmin: Günlük kayıp limiti aşıldı")
        return
    
    # Sinyal gönder
    next_game_num = get_next_game_number(game_info['game_number'])
    await send_new_signal(next_game_num, tahmin_renk, tahmin_sebep, game_info)
    
    # İstatistik güncelle
    performance_stats['excel_stats']['total'] += 1

# EXCEL SİNYAL GÜNCELLEME
async def update_excel_signal(tracker_info, result_type, current_step=None):
    """Excel sinyalini güncelle"""
    try:
        if result_type == 'win':
            performance_stats['excel_stats']['wins'] += 1
            performance_stats['excel_stats']['profit'] += 1
        elif result_type == 'loss':
            performance_stats['excel_stats']['losses'] += 1
            performance_stats['excel_stats']['profit'] -= (2**current_step - 1) if current_step else 1
    except Exception as e:
        print(f"❌ Excel istatistik güncelleme hatası: {e}")

# DİĞER GEREKLİ FONKSİYONLAR
def update_performance_stats(result_type, steps=0, c2_3_type=None, pattern_type=None):
    today = datetime.now(GMT3).strftime('%Y-%m-%d')
    week = datetime.now(GMT3).strftime('%Y-%W')
    
    performance_stats['total_signals'] += 1
    performance_stats['signal_history'].append({
        'timestamp': datetime.now(GMT3),
        'result': result_type,
        'steps': steps,
        'c2_3_type': c2_3_type,
        'pattern_type': pattern_type
    })
    
    if result_type == 'win':
        performance_stats['win_signals'] += 1
        performance_stats['current_streak'] += 1
        performance_stats['max_streak'] = max(performance_stats['max_streak'], performance_stats['current_streak'])
        performance_stats['total_profit'] += 1
        performance_stats['daily_stats'][today]['wins'] += 1
        performance_stats['daily_stats'][today]['profit'] += 1
        performance_stats['weekly_stats'][week]['wins'] += 1
        performance_stats['weekly_stats'][week]['profit'] += 1
    else:
        performance_stats['loss_signals'] += 1
        performance_stats['current_streak'] = 0
        performance_stats['total_profit'] -= (2**steps - 1)
        performance_stats['daily_stats'][today]['losses'] += 1
        performance_stats['daily_stats'][today]['profit'] -= (2**steps - 1)
        performance_stats['weekly_stats'][week]['losses'] += 1
        performance_stats['weekly_stats'][week]['profit'] -= (2**steps - 1)
    
    performance_stats['daily_stats'][today]['signals'] += 1
    performance_stats['weekly_stats'][week]['signals'] += 1

async def send_new_signal(game_num, signal_suit, reason, c2_3_info=None):
    global is_signal_active, daily_signal_count
    try:
        suit_display = get_suit_display_name(signal_suit)
        if c2_3_info:
            c2_3_type, c2_3_desc = c2_3_info.get('c2_3_type', ''), c2_3_info.get('c2_3_description', '')
            trigger_info = f"{c2_3_desc} {c2_3_type}"
        else: 
            c2_3_type, c2_3_desc = '#C2_3', 'KLASİK'
            trigger_info = "KLASİK #C2_3"
        
        # Excel tahminleri için özel işlem
        if "EXCEL" in reason:
            trigger_info = "EXCEL TAHMİN"
            c2_3_type = "#EXCEL"
        
        gmt3_time = datetime.now(GMT3).strftime('%H:%M:%S')
        signal_text = f"🎯 **SİNYAL BAŞLADI** 🎯\n#N{game_num} - {suit_display}\n📊 Tetikleyici: {trigger_info}\n🎯 Sebep: {reason}\n⚡ Strateji: Martingale {MAX_MARTINGALE_STEPS} Seviye\n🕒 {gmt3_time} (GMT+3)\n🔴 SONUÇ: BEKLENİYOR..."
        sent_message = await client.send_message(KANAL_HEDEF, signal_text)
        print(f"🚀 Sinyal gönderildi: #N{game_num} - {suit_display} - {trigger_info}")
        daily_signal_count += 1
        martingale_trackers[game_num] = {
            'message_obj': sent_message, 
            'step': 0, 
            'signal_suit': signal_suit, 
            'sent_game_number': game_num, 
            'expected_game_number_for_check': game_num, 
            'start_time': datetime.now(GMT3), 
            'reason': reason, 
            'c2_3_type': c2_3_type,
            'c2_3_description': c2_3_desc,
            'results': [],
            'is_excel': "EXCEL" in reason  # Excel sinyali mi?
        }
        is_signal_active = True
        
        # Excel'e sinyali kaydet
        game_info = {
            'game_number': game_num,
            'player_cards': '',
            'banker_cards': ''
        }
        save_to_excel(game_info, signal_suit, "SİNYAL_VERİLDİ")
        
    except Exception as e: 
        print(f"❌ Sinyal gönderme hatası: {e}")

async def update_signal_message(tracker_info, result_type, current_step=None, result_details=None):
    try:
        signal_game_num, signal_suit = tracker_info['sent_game_number'], tracker_info['signal_suit']
        suit_display, message_obj, reason = get_suit_display_name(signal_suit), tracker_info['message_obj'], tracker_info.get('reason', '')
        c2_3_type = tracker_info.get('c2_3_type', '#C2_3')
        is_excel_signal = tracker_info.get('is_excel', False)
        
        pattern_type = None
        for pattern in pattern_stats.keys():
            if pattern in reason:
                pattern_type = pattern
                break
        
        duration = datetime.now(GMT3) - tracker_info['start_time']
        duration_str = f"{duration.seconds // 60}d {duration.seconds % 60}s"
        gmt3_time = datetime.now(GMT3).strftime('%H:%M:%S')
        
        if result_details: 
            tracker_info['results'].append(result_details)
        
        if result_type == 'win':
            new_text = f"✅ **KAZANÇ** ✅\n#N{signal_game_num} - {suit_display}\n📊 Sebep: {reason}\n🎯 Seviye: {current_step if current_step else 0}. Seviye\n⏱️ Süre: {duration_str}\n🕒 Bitiş: {gmt3_time}\n🏆 **SONUÇ: KAZANDINIZ!**"
            update_performance_stats('win', current_step if current_step else 0, c2_3_type, pattern_type)
            if is_excel_signal:
                await update_excel_signal(tracker_info, 'win', current_step)
        elif result_type == 'loss':
            new_text = f"❌ **KAYIP** ❌\n#N{signal_game_num} - {suit_display}\n📊 Sebep: {reason}\n🎯 Seviye: {current_step if current_step else MAX_MARTINGALE_STEPS}. Seviye\n⏱️ Süre: {duration_str}\n🕒 Bitiş: {gmt3_time}\n💔 **SONUÇ: KAYBETTİNİZ**"
            update_performance_stats('loss', current_step if current_step else MAX_MARTINGALE_STEPS, c2_3_type, pattern_type)
            if is_excel_signal:
                await update_excel_signal(tracker_info, 'loss', current_step)
        elif result_type == 'progress':
            step_details = f"{current_step}. seviye → #{tracker_info['expected_game_number_for_check']}"
            results_history = "\n".join([f"• {r}" for r in tracker_info['results']]) if tracker_info['results'] else "• İlk deneme"
            new_text = f"🔄 **MARTINGALE İLERLİYOR** 🔄\n#N{signal_game_num} - {suit_display}\n📊 Sebep: {reason}\n🎯 Adım: {step_details}\n⏱️ Süre: {duration_str}\n🕒 Son Güncelleme: {gmt3_time}\n📈 Geçmiş:\n{results_history}\n🎲 **SONRAKİ: #{tracker_info['expected_game_number_for_check']}**"
        elif result_type == 'step_result':
            new_text = f"📊 **ADIM SONUCU** 📊\n#N{signal_game_num} - {suit_display}\n🎯 Adım: {current_step}. seviye\n📋 Sonuç: {result_details}\n⏱️ Süre: {duration_str}\n🕒 Zaman: {gmt3_time}\n🔄 **DEVAM EDİYOR...**"
        
        await message_obj.edit(new_text)
        print(f"✏️ Sinyal güncellendi: #{signal_game_num} - {result_type}")
    except MessageNotModifiedError: 
        pass
    except Exception as e: 
        print(f"❌ Mesaj düzenleme hatası: {e}")

async def check_martingale_trackers():
    global martingale_trackers, is_signal_active
    trackers_to_remove = []
    for signal_game_num, tracker_info in list(martingale_trackers.items()):
        current_step, signal_suit, game_to_check = tracker_info['step'], tracker_info['signal_suit'], tracker_info['expected_game_number_for_check']
        if game_to_check not in game_results: continue
        result_info = game_results.get(game_to_check)
        if not result_info['is_final']: continue
        player_cards_str = result_info['player_cards']
        signal_won_this_step = bool(re.search(re.escape(signal_suit), player_cards_str))
        print(f"🔍 Sinyal kontrol: #{signal_game_num} (Seviye {current_step}) → #{game_to_check}")
        if signal_won_this_step:
            result_details = f"#{game_to_check} ✅ Kazanç - {current_step}. seviye"
            await update_signal_message(tracker_info, 'step_result', current_step, result_details)
            await asyncio.sleep(1)
            await update_signal_message(tracker_info, 'win', current_step)
            trackers_to_remove.append(signal_game_num)
            is_signal_active = False
            recent_games.append({'kazanç': True, 'adim': current_step})
            if len(recent_games) > 20: recent_games.pop(0)
            print(f"🎉 Sinyal #{signal_game_num} KAZANDI! Seviye: {current_step}")
        else:
            result_details = f"#{game_to_check} ❌ Kayıp - {current_step}. seviye"
            await update_signal_message(tracker_info, 'step_result', current_step, result_details)
            await asyncio.sleep(1)
            if current_step < MAX_MARTINGALE_STEPS:
                next_step, next_game_num = current_step + 1, get_next_game_number(game_to_check)
                martingale_trackers[signal_game_num]['step'], martingale_trackers[signal_game_num]['expected_game_number_for_check'] = next_step, next_game_num
                await update_signal_message(tracker_info, 'progress', next_step)
                print(f"📈 Sinyal #{signal_game_num} → {next_step}. seviye → #{next_game_num}")
            else:
                await update_signal_message(tracker_info, 'loss', current_step)
                trackers_to_remove.append(signal_game_num)
                is_signal_active = False
                recent_games.append({'kazanç': False, 'adim': current_step})
                if len(recent_games) > 20: recent_games.pop(0)
                print(f"💔 Sinyal #{signal_game_num} KAYBETTİ! Son seviye: {current_step}")
    for game_num_to_remove in trackers_to_remove:
        if game_num_to_remove in martingale_trackers: 
            del martingale_trackers[game_num_to_remove]

# DİĞER SİSTEM FONKSİYONLARI (kısaltılmış)
async def normal_hibrit_sistemi(game_info):
    trigger_game_num, c2_3_info = game_info['game_number'], {'c2_3_type': game_info.get('c2_3_type'), 'c2_3_description': game_info.get('c2_3_description')}
    print(f"🎯 Normal Hibrit analiz ediyor {c2_3_info['c2_3_description']}...")
    signal_color, reason = analyze_simple_pattern(game_info['player_cards'], game_info['banker_cards'], trigger_game_num)
    if signal_color:
        next_game_num = get_next_game_number(trigger_game_num)
        await send_new_signal(next_game_num, signal_color, reason, c2_3_info)
        print(f"🚀 Normal Hibrit sinyal gönderildi: #{next_game_num} - {reason}")
    else: print(f"🚫 Normal Hibrit: Sinyal yok - {reason}")

async def super_hibrit_sistemi(game_info):
    trigger_game_num, c2_3_info = game_info['game_number'], {'c2_3_type': game_info.get('c2_3_type'), 'c2_3_description': game_info.get('c2_3_description')}
    print(f"🚀 Süper Hibrit analiz ediyor {c2_3_info['c2_3_description']}...")
    signal_color, onay_sebep = besli_onay_sistemi(game_info['player_cards'], game_info['banker_cards'], game_info['game_number'])
    if not signal_color: return print(f"🚫 5'li onay reddedildi: {onay_sebep}")
    filtre_sonuc, filtre_sebep = super_filtre_kontrol(signal_color, onay_sebep, game_info['game_number'])
    if not filtre_sonuc: return print(f"🚫 Süper filtre reddetti: {filtre_sebep}")
    risk_seviye, risk_uyarilar = super_risk_analizi()
    if risk_seviye == "🔴 YÜKSEK RİSK": return print(f"🚫 Yüksek risk: {risk_uyarilar}")
    next_game_num = get_next_game_number(trigger_game_num)
    await send_new_signal(next_game_num, signal_color, f"🚀 SÜPER HİBRİT - {onay_sebep}", c2_3_info)
    print(f"🎯 SÜPER HİBRİT sinyal gönderildi: #{next_game_num}")

# MESAJ İŞLEME
@client.on(events.NewMessage(chats=KANAL_KAYNAK_ID))
@client.on(events.MessageEdited(chats=KANAL_KAYNAK_ID))
async def handle_source_channel_message(event):
    try:
        message = event.message
        text = message.text or ""
        cleaned_text = re.sub(r'\*\*', '', text).strip()
        cleaned_text = re.sub(r'\s+', ' ', cleaned_text).strip()
        gmt3_time = datetime.now(GMT3).strftime('%H:%M:%S')
        print(f"[{gmt3_time}] 📥 Mesaj: #{len(cleaned_text)} karakter")
        game_info = extract_game_info_from_message(cleaned_text)
        if game_info['game_number'] is None: return
        
        # Excel'e kaydet (final sonuçlar için)
        if game_info['is_final']:
            save_to_excel(game_info)
        
        game_results[game_info['game_number']] = game_info
        await check_martingale_trackers()
        if not is_signal_active:
            if game_info['is_final'] and game_info.get('is_c2_3'):
                trigger_game_num, c2_3_type, c2_3_desc = game_info['game_number'], game_info['c2_3_type'], game_info['c2_3_description']
                print(f"🎯 {c2_3_desc} tespit edildi: #{trigger_game_num} - {c2_3_type}")
                
                if SISTEM_MODU == "normal_hibrit": 
                    await normal_hibrit_sistemi(game_info)
                elif SISTEM_MODU == "super_hibrit": 
                    await super_hibrit_sistemi(game_info)
                elif SISTEM_MODU == "quantum_hibrit":
                    await quantum_hibrit_sistemi(game_info)
                elif SISTEM_MODU == "quantum_pro":
                    await quantum_pro_sistemi(game_info)
                elif SISTEM_MODU == "master_elite":
                    await master_elite_sistemi(game_info)
                elif SISTEM_MODU == "excel_tahmin":
                    await excel_tahmin_sistemi(game_info)
                    
    except Exception as e: print(f"❌ Mesaj işleme hatası: {e}")

# YENİ KOMUTLAR
@client.on(events.NewMessage(pattern='(?i)/mod_excel'))
async def handle_mod_excel(event):
    global SISTEM_MODU
    SISTEM_MODU = "excel_tahmin"
    await event.reply("📊 EXCEL TAHMİN modu aktif! Excel veri analizi + 3 martingale.")

@client.on(events.NewMessage(pattern='(?i)/excel_durum'))
async def handle_excel_durum(event):
    analysis = get_excel_performance()
    await event.reply(analysis)

@client.on(events.NewMessage(pattern='(?i)/excel_analiz'))
async def handle_excel_analiz(event):
    tahmin_renk, tahmin_sebep = analyze_excel_pattern()
    if tahmin_renk:
        await event.reply(f"📊 **EXCEL ANALİZ SONUCU**\n\n🎯 Tahmin: {get_suit_display_name(tahmin_renk)}\n📈 Sebep: {tahmin_sebep}")
    else:
        await event.reply(f"📊 **EXCEL ANALİZ SONUCU**\n\n❌ {tahmin_sebep}")

@client.on(events.NewMessage(pattern='(?i)/excel_temizle'))
async def handle_excel_temizle(event):
    if event.sender_id != ADMIN_ID: 
        return await event.reply("❌ Yetkiniz yok!")
    try:
        if os.path.exists(EXCEL_FILE):
            os.remove(EXCEL_FILE)
        init_excel_file()
        await event.reply("✅ Excel verileri temizlendi! Yeni dosya oluşturuldu.")
    except Exception as e:
        await event.reply(f"❌ Excel temizleme hatası: {e}")

# DİĞER KOMUTLAR (kısaltılmış)
@client.on(events.NewMessage(pattern='(?i)/basla'))
async def handle_start(event): 
    await event.reply("🤖 Royal Baccarat Bot Aktif! 🎯")

@client.on(events.NewMessage(pattern='(?i)/durum'))
async def handle_durum(event):
    aktif_sinyal = "✅ Evet" if is_signal_active else "❌ Hayır"
    gmt3_time = datetime.now(GMT3).strftime('%H:%M:%S')
    aktif_takipciler = "\n".join([f"• #{num} (Seviye {info['step']})" for num, info in martingale_trackers.items()])
    if not aktif_takipciler: 
        aktif_takipciler = "• Aktif sinyal yok"
    
    durum_mesaji = f"""🤖 **ROYAL BACCARAT BOT** 🤖

🟢 **Durum:** Çalışıyor
🎯 **Aktif Sinyal:** {aktif_sinyal}
📊 **Aktif Takipçiler:**
{aktif_takipciler}
📈 **Trend:** {color_trend[-5:] if color_trend else 'Yok'}
🎛️ **Mod:** {SISTEM_MODU}
🕒 **Saat:** {gmt3_time} (GMT+3)
📨 **Günlük Sinyal:** {daily_signal_count}

⚡ **Sistem:** Hibrit Pattern + Martingale {MAX_MARTINGALE_STEPS} Seviye
"""
    await event.reply(durum_mesaji)

# DOSYA BAŞLANGICI
if __name__ == '__main__':
    print("🤖 ROYAL BACCARAT BOT BAŞLIYOR...")
    print(f"🔧 API_ID: {API_ID}")
    print(f"🎯 Kaynak Kanal: {KANAL_KAYNAK_ID}")
    print(f"📤 Hedef Kanal: {KANAL_HEDEF}")
    print(f"👤 Admin ID: {ADMIN_ID}")
    print(f"🎛️ Varsayılan Mod: {SISTEM_MODU}")
    print(f"📊 C2-3 Analiz Sistemi: AKTİF")
    print(f"📈 Pattern Performans Takibi: AKTİF")
    print(f"📊 Excel Tahmin Sistemi: AKTİF")
    print(f"🕒 Saat Dilimi: GMT+3")
    
    # Excel dosyasını başlat
    init_excel_file()
    
    print("⏳ Bağlanıyor...")
    try:
        with client: 
            client.run_until_disconnected()
    except KeyboardInterrupt: 
        print("\n👋 Bot durduruluyor...")
    except Exception as e: 
        print(f"❌ Bot başlangıç hatası: {e}")