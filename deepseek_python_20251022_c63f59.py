# -*- coding: utf-8 -*-
import re, json, os, asyncio, sys, pytz
from datetime import datetime, timedelta
from telethon import TelegramClient, events
from telethon.errors import FloodWaitError, MessageNotModifiedError
from collections import defaultdict, deque
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import random

API_ID = 29581698
API_HASH = '0caabd4263f1d4e5f753659a787c2e7d'
BOT_TOKEN = ''  # 🔑 Buraya bot tokenınızı yazın
KANAL_KAYNAK_ID = -1001626824569
KANAL_HEDEF = "@royalbaccfree"  # 📢 Hedef kanal
ADMIN_ID = 1136442929  # 👑 Admin ID
SISTEM_MODU = "normal_hibrit"
GMT3 = pytz.timezone('Europe/Istanbul')
client = TelegramClient('royal_bot', API_ID, API_HASH).start(bot_token=BOT_TOKEN)

game_results, martingale_trackers, color_trend, recent_games = {}, {}, [], []
MAX_MARTINGALE_STEPS, MAX_GAME_NUMBER, is_signal_active, daily_signal_count = 3, 1440, False, 0

# Excel dosyası için değişkenler
EXCEL_FILE = "baccarat_data.xlsx"
excel_data = []

# Güncellenmiş C2_3 istatistik yapısı
C2_3_TYPES = {
    '#C2_3': {'emoji': '🔴', 'name': 'KLASİK', 'confidence': 0.9, 'stats': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0}},
    '#C2_2': {'emoji': '🔵', 'name': 'ALTERNATİF', 'confidence': 0.7, 'stats': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0}},
    '#C3_2': {'emoji': '🟢', 'name': 'VARYANT', 'confidence': 0.6, 'stats': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0}},
    '#C3_3': {'emoji': '🟡', 'name': 'ÖZEL', 'confidence': 0.7, 'stats': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0}}
}

# İstatistik veri yapıları
performance_stats = {
    'total_signals': 0,
    'win_signals': 0,
    'loss_signals': 0,
    'total_profit': 0,
    'current_streak': 0,
    'max_streak': 0,
    'daily_stats': defaultdict(lambda: {'signals': 0, 'wins': 0, 'losses': 0, 'profit': 0}),
    'weekly_stats': defaultdict(lambda: {'signals': 0, 'wins': 0, 'losses': 0, 'profit': 0}),
    'signal_history': deque(maxlen=1000),
    'c2_3_performance': C2_3_TYPES.copy(),
    'excel_stats': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0}
}

# PATTERN İSTATİSTİKLERİ
pattern_stats = {
    '🎯 GÜÇLÜ EL': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0, 'avg_steps': 0},
    '🏆 DOĞAL KAZANÇ': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0, 'avg_steps': 0},
    '📊 5+ KART': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0, 'avg_steps': 0},
    '🚨 3x TEKRAR': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0, 'avg_steps': 0},
    '📈 STANDART SİNYAL': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0, 'avg_steps': 0},
    '✅ 5-Lİ ONAY': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0, 'avg_steps': 0},
    '🚀 SÜPER HİBRİT': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0, 'avg_steps': 0},
    '🎯 KLASİK #C2_3': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0, 'avg_steps': 0},
    '📊 EXCEL TAHMİN': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0, 'avg_steps': 0}
}

# EXCEL ANALİZ FONKSİYONLARI
def init_excel_file():
    """Excel dosyasını başlat"""
    try:
        if not os.path.exists(EXCEL_FILE):
            wb = Workbook()
            ws = wb.active
            ws.title = "Baccarat Data"
            # Başlıklar
            headers = [
                "Tarih", "Saat", "Oyun No", "Oyuncu Kartları", "Banker Kartları", 
                "Oyuncu Toplam", "Banker Toplam", "Kazanan", "Sinyal Rengi",
                "Tahmin Rengi", "Sonuç", "Martingale Seviye", "Kazanç/Kayıp"
            ]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            wb.save(EXCEL_FILE)
            print("✅ Excel dosyası oluşturuldu")
        else:
            print("✅ Excel dosyası zaten mevcut")
    except Exception as e:
        print(f"❌ Excel dosyası oluşturma hatası: {e}")

def save_to_excel(game_info, signal_color=None, result=None, martingale_step=0, profit=0):
    """Excel'e veri kaydet"""
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        # Son satırı bul
        row = ws.max_row + 1
        
        # Tarih ve saat
        now = datetime.now(GMT3)
        date_str = now.strftime('%Y-%m-%d')
        time_str = now.strftime('%H:%M:%S')
        
        # Oyuncu ve banker toplamını hesapla
        player_total = calculate_hand_total(game_info['player_cards'])
        banker_total = calculate_hand_total(game_info['banker_cards'])
        
        # Kazananı belirle
        winner = determine_winner(player_total, banker_total)
        
        # Tahmin rengi (eğer sinyal varsa)
        predicted_color = signal_color if signal_color else "YOK"
        
        # Sonuç
        result_text = result if result else "BEKLENİYOR"
        
        # Verileri yaz
        data = [
            date_str, time_str, game_info['game_number'], 
            game_info['player_cards'], game_info['banker_cards'],
            player_total, banker_total, winner, 
            extract_largest_value_suit(game_info['player_cards']),
            predicted_color, result_text, martingale_step, profit
        ]
        
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
        
        wb.save(EXCEL_FILE)
        print(f"✅ Excel'e kaydedildi: #{game_info['game_number']}")
        
    except Exception as e:
        print(f"❌ Excel kayıt hatası: {e}")

def calculate_hand_total(cards_str):
    """Kart toplamını hesapla"""
    try:
        cards = re.findall(r'(10|[A2-9TJQK])([♣♦♥♠])', cards_str)
        total = 0
        for card_char, suit in cards:
            value = get_baccarat_value(card_char)
            total += value
        return total % 10
    except:
        return 0

def determine_winner(player_total, banker_total):
    """Kazananı belirle"""
    if player_total > banker_total:
        return "PLAYER"
    elif banker_total > player_total:
        return "BANKER"
    else:
        return "TIE"

def analyze_excel_pattern():
    """Excel verilerine göre tahmin yap"""
    try:
        if not os.path.exists(EXCEL_FILE):
            return None, "Excel dosyası bulunamadı"
        
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        if ws.max_row <= 1:
            return None, "Yeterli veri yok"
        
        # Son 50 kaydı analiz et
        data = []
        for row in range(max(2, ws.max_row - 49), ws.max_row + 1):
            signal_color = ws.cell(row=row, column=9).value  # Sinyal Rengi sütunu
            result = ws.cell(row=row, column=11).value  # Sonuç sütunu
            
            if signal_color and result in ['KAZANÇ', 'KAYIP']:
                data.append({
                    'color': signal_color,
                    'win': result == 'KAZANÇ'
                })
        
        if len(data) < 10:
            return None, f"Yeterli veri yok ({len(data)} kayıt)"
        
        # Renk analizi
        color_stats = {}
        for suit in ['♠', '♥', '♦', '♣']:
            color_data = [d for d in data if d['color'] == suit]
            if color_data:
                wins = sum(1 for d in color_data if d['win'])
                total = len(color_data)
                win_rate = wins / total if total > 0 else 0
                color_stats[suit] = {
                    'total': total,
                    'wins': wins,
                    'win_rate': win_rate,
                    'recent_count': sum(1 for d in data[-10:] if d['color'] == suit)
                }
        
        if not color_stats:
            return None, "Renk verisi bulunamadı"
        
        # En iyi rengi seç (kazanç oranı + frekans)
        best_color = None
        best_score = -1
        
        for suit, stats in color_stats.items():
            # Skor: kazanç oranı * 0.7 + son 10'daki frekans * 0.3
            score = (stats['win_rate'] * 0.7) + ((stats['recent_count'] / 10) * 0.3)
            
            if score > best_score and stats['total'] >= 3:
                best_score = score
                best_color = suit
        
        if best_color and best_score >= 0.4:
            win_rate = color_stats[best_color]['win_rate'] * 100
            return best_color, f"📊 EXCEL TAHMİN (%{win_rate:.1f} başarı)"
        
        return None, f"Yeterli güven yok (en iyi skor: {best_score:.2f})"
        
    except Exception as e:
        return None, f"Analiz hatası: {e}"

def get_excel_performance():
    """Excel tahmin performansını getir"""
    stats = performance_stats['excel_stats']
    if stats['total'] == 0:
        return "📊 Excel Tahmin: Henüz veri yok"
    
    win_rate = (stats['wins'] / stats['total']) * 100
    return f"""📊 **EXCEL TAHMİN PERFORMANSI**

• Toplam Sinyal: {stats['total']}
• Kazanç: {stats['wins']} | Kayıp: {stats['losses']}
• Başarı Oranı: %{win_rate:.1f}
• Toplam Kâr: {stats['profit']} birim
"""

# EXCEL TAHMİN SİSTEMİ
async def excel_tahmin_sistemi(game_info):
    """Excel tabanlı tahmin sistemi"""
    print("📊 EXCEL TAHMİN analiz başlıyor...")
    
    # Excel'e mevcut oyunu kaydet (sonuçsuz)
    save_to_excel(game_info)
    
    # Excel analizi yap
    tahmin_renk, tahmin_sebep = analyze_excel_pattern()
    
    if not tahmin_renk:
        print(f"🚫 Excel Tahmin: {tahmin_sebep}")
        return
    
    # Risk kontrolü
    risk_seviye, risk_uyarilar = super_risk_analizi()
    if risk_seviye == "🔴 YÜKSEK RİSK":
        print(f"🚫 Excel Tahmin: Yüksek risk - {risk_uyarilar}")
        return
    
    # Günlük performans kontrolü
    daily = get_daily_stats()
    if daily['profit'] <= -10:
        print("🚫 Excel Tahmin: Günlük kayıp limiti aşıldı")
        return
    
    # Sinyal gönder
    next_game_num = get_next_game_number(game_info['game_number'])
    await send_new_signal(next_game_num, tahmin_renk, tahmin_sebep, game_info)
    
    # İstatistik güncelle
    performance_stats['excel_stats']['total'] += 1

# EXCEL SİNYAL GÜNCELLEME
async def update_excel_signal(tracker_info, result_type, current_step=None):
    """Excel sinyalini güncelle"""
    try:
        if result_type == 'win':
            performance_stats['excel_stats']['wins'] += 1
            performance_stats['excel_stats']['profit'] += 1
        elif result_type == 'loss':
            performance_stats['excel_stats']['losses'] += 1
            performance_stats['excel_stats']['profit'] -= (2**current_step - 1) if current_step else 1
    except Exception as e:
        print(f"❌ Excel istatistik güncelleme hatası: {e}")

# GÜNCELLENMİŞ SİNYAL GÖNDERME
async def send_new_signal(game_num, signal_suit, reason, c2_3_info=None):
    global is_signal_active, daily_signal_count
    try:
        suit_display = get_suit_display_name(signal_suit)
        if c2_3_info:
            c2_3_type, c2_3_desc = c2_3_info.get('c2_3_type', ''), c2_3_info.get('c2_3_description', '')
            trigger_info = f"{c2_3_desc} {c2_3_type}"
        else: 
            c2_3_type, c2_3_desc = '#C2_3', 'KLASİK'
            trigger_info = "KLASİK #C2_3"
        
        # Excel tahminleri için özel işlem
        if "EXCEL" in reason:
            trigger_info = "EXCEL TAHMİN"
            c2_3_type = "#EXCEL"
        
        gmt3_time = datetime.now(GMT3).strftime('%H:%M:%S')
        signal_text = f"🎯 **SİNYAL BAŞLADI** 🎯\n#N{game_num} - {suit_display}\n📊 Tetikleyici: {trigger_info}\n🎯 Sebep: {reason}\n⚡ Strateji: Martingale {MAX_MARTINGALE_STEPS} Seviye\n🕒 {gmt3_time} (GMT+3)\n🔴 SONUÇ: BEKLENİYOR..."
        sent_message = await client.send_message(KANAL_HEDEF, signal_text)
        print(f"🚀 Sinyal gönderildi: #N{game_num} - {suit_display} - {trigger_info}")
        daily_signal_count += 1
        martingale_trackers[game_num] = {
            'message_obj': sent_message, 
            'step': 0, 
            'signal_suit': signal_suit, 
            'sent_game_number': game_num, 
            'expected_game_number_for_check': game_num, 
            'start_time': datetime.now(GMT3), 
            'reason': reason, 
            'c2_3_type': c2_3_type,
            'c2_3_description': c2_3_desc,
            'results': [],
            'is_excel': "EXCEL" in reason  # Excel sinyali mi?
        }
        is_signal_active = True
        
        # Excel'e sinyali kaydet
        game_info = {
            'game_number': game_num,
            'player_cards': '',
            'banker_cards': ''
        }
        save_to_excel(game_info, signal_suit, "SİNYAL_VERİLDİ")
        
    except Exception as e: 
        print(f"❌ Sinyal gönderme hatası: {e}")

# GÜNCELLENMİŞ SİNYAL GÜNCELLEME
async def update_signal_message(tracker_info, result_type, current_step=None, result_details=None):
    try:
        signal_game_num, signal_suit = tracker_info['sent_game_number'], tracker_info['signal_suit']
        suit_display, message_obj, reason = get_suit_display_name(signal_suit), tracker_info['message_obj'], tracker_info.get('reason', '')
        c2_3_type = tracker_info.get('c2_3_type', '#C2_3')
        is_excel_signal = tracker_info.get('is_excel', False)
        
        pattern_type = None
        for pattern in pattern_stats.keys():
            if pattern in reason:
                pattern_type = pattern
                break
        
        duration = datetime.now(GMT3) - tracker_info['start_time']
        duration_str = f"{duration.seconds // 60}d {duration.seconds % 60}s"
        gmt3_time = datetime.now(GMT3).strftime('%H:%M:%S')
        
        if result_details: 
            tracker_info['results'].append(result_details)
        
        if result_type == 'win':
            new_text = f"✅ **KAZANÇ** ✅\n#N{signal_game_num} - {suit_display}\n📊 Sebep: {reason}\n🎯 Seviye: {current_step if current_step else 0}. Seviye\n⏱️ Süre: {duration_str}\n🕒 Bitiş: {gmt3_time}\n🏆 **SONUÇ: KAZANDINIZ!**"
            update_performance_stats('win', current_step if current_step else 0, c2_3_type, pattern_type)
            if is_excel_signal:
                await update_excel_signal(tracker_info, 'win', current_step)
        elif result_type == 'loss':
            new_text = f"❌ **KAYIP** ❌\n#N{signal_game_num} - {suit_display}\n📊 Sebep: {reason}\n🎯 Seviye: {current_step if current_step else MAX_MARTINGALE_STEPS}. Seviye\n⏱️ Süre: {duration_str}\n🕒 Bitiş: {gmt3_time}\n💔 **SONUÇ: KAYBETTİNİZ**"
            update_performance_stats('loss', current_step if current_step else MAX_MARTINGALE_STEPS, c2_3_type, pattern_type)
            if is_excel_signal:
                await update_excel_signal(tracker_info, 'loss', current_step)
        elif result_type == 'progress':
            step_details = f"{current_step}. seviye → #{tracker_info['expected_game_number_for_check']}"
            results_history = "\n".join([f"• {r}" for r in tracker_info['results']]) if tracker_info['results'] else "• İlk deneme"
            new_text = f"🔄 **MARTINGALE İLERLİYOR** 🔄\n#N{signal_game_num} - {suit_display}\n📊 Sebep: {reason}\n🎯 Adım: {step_details}\n⏱️ Süre: {duration_str}\n🕒 Son Güncelleme: {gmt3_time}\n📈 Geçmiş:\n{results_history}\n🎲 **SONRAKİ: #{tracker_info['expected_game_number_for_check']}**"
        elif result_type == 'step_result':
            new_text = f"📊 **ADIM SONUCU** 📊\n#N{signal_game_num} - {suit_display}\n🎯 Adım: {current_step}. seviye\n📋 Sonuç: {result_details}\n⏱️ Süre: {duration_str}\n🕒 Zaman: {gmt3_time}\n🔄 **DEVAM EDİYOR...**"
        
        await message_obj.edit(new_text)
        print(f"✏️ Sinyal güncellendi: #{signal_game_num} - {result_type}")
    except MessageNotModifiedError: 
        pass
    except Exception as e: 
        print(f"❌ Mesaj düzenleme hatası: {e}")

# GÜNCELLENMİŞ MESAJ İŞLEME
@client.on(events.NewMessage(chats=KANAL_KAYNAK_ID))
@client.on(events.MessageEdited(chats=KANAL_KAYNAK_ID))
async def handle_source_channel_message(event):
    try:
        message = event.message
        text = message.text or ""
        cleaned_text = re.sub(r'\*\*', '', text).strip()
        cleaned_text = re.sub(r'\s+', ' ', cleaned_text).strip()
        gmt3_time = datetime.now(GMT3).strftime('%H:%M:%S')
        print(f"[{gmt3_time}] 📥 Mesaj: #{len(cleaned_text)} karakter")
        game_info = extract_game_info_from_message(cleaned_text)
        if game_info['game_number'] is None: return
        
        # Excel'e kaydet (final sonuçlar için)
        if game_info['is_final']:
            save_to_excel(game_info)
        
        game_results[game_info['game_number']] = game_info
        await check_martingale_trackers()
        if not is_signal_active:
            if game_info['is_final'] and game_info.get('is_c2_3'):
                trigger_game_num, c2_3_type, c2_3_desc = game_info['game_number'], game_info['c2_3_type'], game_info['c2_3_description']
                print(f"🎯 {c2_3_desc} tespit edildi: #{trigger_game_num} - {c2_3_type}")
                
                if SISTEM_MODU == "normal_hibrit": 
                    await normal_hibrit_sistemi(game_info)
                elif SISTEM_MODU == "super_hibrit": 
                    await super_hibrit_sistemi(game_info)
                elif SISTEM_MODU == "quantum_hibrit":
                    await quantum_hibrit_sistemi(game_info)
                elif SISTEM_MODU == "quantum_pro":
                    await quantum_pro_sistemi(game_info)
                elif SISTEM_MODU == "master_elite":
                    await master_elite_sistemi(game_info)
                elif SISTEM_MODU == "excel_tahmin":
                    await excel_tahmin_sistemi(game_info)
                    
    except Exception as e: print(f"❌ Mesaj işleme hatası: {e}")

# YENİ KOMUTLAR
@client.on(events.NewMessage(pattern='(?i)/mod_excel'))
async def handle_mod_excel(event):
    global SISTEM_MODU
    SISTEM_MODU = "excel_tahmin"
    await event.reply("📊 EXCEL TAHMİN modu aktif! Excel veri analizi + 3 martingale.")

@client.on(events.NewMessage(pattern='(?i)/excel_durum'))
async def handle_excel_durum(event):
    analysis = get_excel_performance()
    await event.reply(analysis)

@client.on(events.NewMessage(pattern='(?i)/excel_analiz'))
async def handle_excel_analiz(event):
    tahmin_renk, tahmin_sebep = analyze_excel_pattern()
    if tahmin_renk:
        await event.reply(f"📊 **EXCEL ANALİZ SONUCU**\n\n🎯 Tahmin: {get_suit_display_name(tahmin_renk)}\n📈 Sebep: {tahmin_sebep}")
    else:
        await event.reply(f"📊 **EXCEL ANALİZ SONUCU**\n\n❌ {tahmin_sebep}")

@client.on(events.NewMessage(pattern='(?i)/excel_temizle'))
async def handle_excel_temizle(event):
    if event.sender_id != ADMIN_ID: 
        return await event.reply("❌ Yetkiniz yok!")
    try:
        if os.path.exists(EXCEL_FILE):
            os.remove(EXCEL_FILE)
        init_excel_file()
        await event.reply("✅ Excel verileri temizlendi! Yeni dosya oluşturuldu.")
    except Exception as e:
        await event.reply(f"❌ Excel temizleme hatası: {e}")

# DOSYA BAŞLANGICI
if __name__ == '__main__':
    print("🤖 ROYAL BACCARAT BOT BAŞLIYOR...")
    print(f"🔧 API_ID: {API_ID}")
    print(f"🎯 Kaynak Kanal: {KANAL_KAYNAK_ID}")
    print(f"📤 Hedef Kanal: {KANAL_HEDEF}")
    print(f"👤 Admin ID: {ADMIN_ID}")
    print(f"🎛️ Varsayılan Mod: {SISTEM_MODU}")
    print(f"📊 C2-3 Analiz Sistemi: AKTİF")
    print(f"📈 Pattern Performans Takibi: AKTİF")
    print(f"⚛️ Quantum Hibrit Sistem: AKTİF")
    print(f"🚀 Quantum PRO Sistem: AKTİF")
    print(f"🏆 Master Elite Sistem: AKTİF")
    print(f"📊 Excel Tahmin Sistemi: AKTİF")
    print(f"🕒 Saat Dilimi: GMT+3")
    
    # Excel dosyasını başlat
    init_excel_file()
    
    print("⏳ Bağlanıyor...")
    try:
        with client: 
            client.run_until_disconnected()
    except KeyboardInterrupt: 
        print("\n👋 Bot durduruluyor...")
    except Exception as e: 
        print(f"❌ Bot başlangıç hatası: {e}")