# -*- coding: utf-8 -*-
import re, json, os, asyncio, sys, pytz, logging
from datetime import datetime, timedelta
from telethon import TelegramClient, events
from telethon.errors import FloodWaitError, MessageNotModifiedError, RPCError
from telethon.tl.types import DocumentAttributeFilename
from collections import defaultdict, deque
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

# Logging ayarı
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('/var/log/kolera_bot.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger('KoleraBot')

API_ID = 29581698
API_HASH = '0caabd4263f1d4e5f753659a787c2e7d'
BOT_TOKEN = ''  # 🔑 Buraya bot tokenınızı yazın
KANAL_KAYNAK_ID = -1001626824569
KANAL_HEDEF = "@royalbaccfree"  # 📢 Hedef kanal
ADMIN_ID = 1136442929  # 👑 Admin ID
SISTEM_MODU = "normal_hibrit"
GMT3 = pytz.timezone('Europe/Istanbul')

# Global değişkenler
game_results, martingale_trackers, color_trend, recent_games = {}, {}, [], []
MAX_MARTINGALE_STEPS, MAX_GAME_NUMBER, is_signal_active, daily_signal_count = 3, 1440, False, 0

# Excel dosyası için renk tanımlamaları
EXCEL_FILE = "/root/0zctn/royal_baccarat_data.xlsx"
RED_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
BLUE_FILL = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
GREEN_FILL = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
BLACK_FILL = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

# Güncellenmiş C2_3 istatistik yapısı
C2_3_TYPES = {
    '#C2_3': {'emoji': '🔴', 'name': 'KLASİK', 'confidence': 0.9, 'stats': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0}},
    '#C2_2': {'emoji': '🔵', 'name': 'ALTERNATİF', 'confidence': 0.7, 'stats': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0}},
    '#C3_2': {'emoji': '🟢', 'name': 'VARYANT', 'confidence': 0.6, 'stats': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0}},
    '#C3_3': {'emoji': '🟡', 'name': 'ÖZEL', 'confidence': 0.7, 'stats': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0}}
}

# İstatistik veri yapıları
performance_stats = {
    'total_signals': 0,
    'win_signals': 0,
    'loss_signals': 0,
    'total_profit': 0,
    'current_streak': 0,
    'max_streak': 0,
    'daily_stats': defaultdict(lambda: {'signals': 0, 'wins': 0, 'losses': 0, 'profit': 0}),
    'weekly_stats': defaultdict(lambda: {'signals': 0, 'wins': 0, 'losses': 0, 'profit': 0}),
    'signal_history': deque(maxlen=1000),
    'c2_3_performance': C2_3_TYPES.copy()
}

# PATTERN İSTATİSTİKLERİ
pattern_stats = {
    '🎯 GÜÇLÜ EL': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0, 'avg_steps': 0},
    '🏆 DOĞAL KAZANÇ': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0, 'avg_steps': 0},
    '📊 5+ KART': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0, 'avg_steps': 0},
    '🚨 3x TEKRAR': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0, 'avg_steps': 0},
    '📈 STANDART SİNYAL': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0, 'avg_steps': 0},
    '✅ 5-Lİ ONAY': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0, 'avg_steps': 0},
    '🚀 SÜPER HİBRİT': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0, 'avg_steps': 0},
    '🎯 KLASİK #C2_3': {'total': 0, 'wins': 0, 'losses': 0, 'profit': 0, 'avg_steps': 0}
}

# Excel dosyasını başlat
def init_excel_file():
    try:
        if not os.path.exists(EXCEL_FILE):
            wb = Workbook()
            ws = wb.active
            ws.title = "Baccarat Data"
            
            # Başlıkları oluştur
            headers = [
                "Oyun No", "Tarih", "Saat", "Player Kartları", "Banker Kartları", 
                "Player Toplam", "Banker Toplam", "Kazanan", "C2 Tipi", "Renk Tahmini",
                "Pattern Tipi", "Sinyal Seviyesi", "Sonuç", "Kazanç/Kayıp", "Toplam Kâr"
            ]
            
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header).font = Font(bold=True)
            
            wb.save(EXCEL_FILE)
            print(f"✅ Excel dosyası oluşturuldu: {EXCEL_FILE}")
        else:
            print(f"✅ Excel dosyası zaten mevcut: {EXCEL_FILE}")
    except Exception as e:
        print(f"❌ Excel dosyası oluşturma hatası: {e}")

# Excel'e veri kaydet
def save_to_excel(game_data):
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        
        # Yeni satır ekle
        row = ws.max_row + 1
        
        # Verileri yaz
        ws.cell(row=row, column=1, value=game_data.get('game_number'))
        ws.cell(row=row, column=2, value=game_data.get('date'))
        ws.cell(row=row, column=3, value=game_data.get('time'))
        ws.cell(row=row, column=4, value=game_data.get('player_cards'))
        ws.cell(row=row, column=5, value=game_data.get('banker_cards'))
        ws.cell(row=row, column=6, value=game_data.get('player_total'))
        ws.cell(row=row, column=7, value=game_data.get('banker_total'))
        ws.cell(row=row, column=8, value=game_data.get('winner'))
        ws.cell(row=row, column=9, value=game_data.get('c2_type'))
        
        # Renk tahmini hücresini renklendir
        color_cell = ws.cell(row=row, column=10, value=game_data.get('color_prediction'))
        color_pred = game_data.get('color_prediction', '')
        if '🔴' in color_pred or 'MAÇA' in color_pred:
            color_cell.fill = RED_FILL
            color_cell.value = "🔴 MAÇA"
        elif '🔵' in color_pred or 'KALP' in color_pred:
            color_cell.fill = BLUE_FILL  
            color_cell.value = "🔵 KALP"
        elif '🟢' in color_pred or 'KARO' in color_pred:
            color_cell.fill = GREEN_FILL
            color_cell.value = "🟢 KARO"
        elif '⚫' in color_pred or 'SİNEK' in color_pred:
            color_cell.fill = BLACK_FILL
            color_cell.value = "⚫ SİNEK"
        
        ws.cell(row=row, column=11, value=game_data.get('pattern_type'))
        ws.cell(row=row, column=12, value=game_data.get('signal_level'))
        
        # Sonuç hücresini renklendir
        result_cell = ws.cell(row=row, column=13, value=game_data.get('result'))
        if game_data.get('result') == 'KAZANÇ':
            result_cell.fill = GREEN_FILL
        elif game_data.get('result') == 'KAYIP':
            result_cell.fill = RED_FILL
            
        ws.cell(row=row, column=14, value=game_data.get('profit_loss'))
        ws.cell(row=row, column=15, value=game_data.get('total_profit'))
        
        wb.save(EXCEL_FILE)
        print(f"✅ Excel'e kaydedildi: #{game_data.get('game_number')}")
        
    except Exception as e:
        print(f"❌ Excel kaydetme hatası: {e}")

# EXCEL KART ANALİZ FONKSİYONLARI
def analyze_cards_by_c2_type():
    """Excel'deki verilere göre C2 tiplerine göre kart analizi yapar"""
    try:
        if not os.path.exists(EXCEL_FILE):
            return "❌ Excel dosyası bulunamadı"

        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active

        if ws.max_row <= 1:
            return "📊 Excel'de henüz veri yok"

        c2_card_stats = {}
        c2_suit_stats = {}
        c2_value_stats = {}

        for row in range(2, ws.max_row + 1):
            c2_type = ws.cell(row=row, column=9).value
            player_cards = ws.cell(row=row, column=4).value
            banker_cards = ws.cell(row=row, column=5).value
            result = ws.cell(row=row, column=13).value

            if c2_type and c2_type != 'N/A' and player_cards and banker_cards:
                if c2_type not in c2_card_stats:
                    c2_card_stats[c2_type] = {
                        'player_cards': [],
                        'banker_cards': [],
                        'total_games': 0,
                        'wins': 0,
                        'losses': 0
                    }
                    c2_suit_stats[c2_type] = {'♠': 0, '♥': 0, '♦': 0, '♣': 0}
                    c2_value_stats[c2_type] = defaultdict(int)

                c2_card_stats[c2_type]['player_cards'].append(player_cards)
                c2_card_stats[c2_type]['banker_cards'].append(banker_cards)
                c2_card_stats[c2_type]['total_games'] += 1

                if result == 'KAZANÇ':
                    c2_card_stats[c2_type]['wins'] += 1
                elif result == 'KAYIP':
                    c2_card_stats[c2_type]['losses'] += 1

                all_cards_text = player_cards + banker_cards
                suits = re.findall(r'[♠♥♦♣]', all_cards_text)
                for suit in suits:
                    c2_suit_stats[c2_type][suit] += 1

                card_values = re.findall(r'(10|[A2-9TJQK])', all_cards_text)
                for value in card_values:
                    c2_value_stats[c2_type][value] += 1

        return generate_card_analysis_report(c2_card_stats, c2_suit_stats, c2_value_stats)

    except Exception as e:
        return f"❌ Kart analiz hatası: {e}"

def generate_card_analysis_report(c2_card_stats, c2_suit_stats, c2_value_stats):
    """Kart analiz raporu oluşturur"""
    if not c2_card_stats:
        return "📊 C2 tipine göre kart verisi bulunamadı"

    report = "🃏 **C2 TİPİNE GÖRE KART ANALİZİ** 🃏\n\n"

    for c2_type, stats in c2_card_stats.items():
        total_games = stats['total_games']
        win_rate = (stats['wins'] / total_games * 100) if total_games > 0 else 0
        
        report += f"🎯 **{c2_type}** - {total_games} oyun (%{win_rate:.1f} başarı)\n"

        suit_stats = c2_suit_stats[c2_type]
        total_suits = sum(suit_stats.values())
        
        report += "🎨 **Renk Dağılımı:**\n"
        for suit, count in suit_stats.items():
            if total_suits > 0:
                percentage = (count / total_suits) * 100
                suit_name = get_suit_display_name(suit)
                report += f"   {suit_name}: {count} (%{percentage:.1f})\n"

        value_stats = c2_value_stats[c2_type]
        total_values = sum(value_stats.values())
        
        report += "🔢 **Kart Değerleri (Top 5):**\n"
        sorted_values = sorted(value_stats.items(), key=lambda x: x[1], reverse=True)[:5]
        for value, count in sorted_values:
            if total_values > 0:
                percentage = (count / total_values) * 100
                report += f"   {value}: {count} (%{percentage:.1f})\n"

        combo_analysis = analyze_card_combinations(stats['player_cards'], stats['banker_cards'])
        report += f"🔄 **Ortalama Kart Sayısı:** {combo_analysis['avg_cards']:.1f}\n"
        
        report += "\n" + "─" * 40 + "\n\n"

    return report

def analyze_card_combinations(player_cards_list, banker_cards_list):
    total_cards = 0
    total_games = len(player_cards_list)
    
    for i in range(total_games):
        player_cards = player_cards_list[i]
        banker_cards = banker_cards_list[i]
        
        player_card_count = len(re.findall(r'(10|[A2-9TJQK])[♠♥♦♣]', player_cards))
        banker_card_count = len(re.findall(r'(10|[A2-9TJQK])[♠♥♦♣]', banker_cards))
        
        total_cards += player_card_count + banker_card_count
    
    return {
        'avg_cards': total_cards / total_games if total_games > 0 else 0
    }

def analyze_winning_cards_patterns():
    try:
        if not os.path.exists(EXCEL_FILE):
            return "❌ Excel dosyası bulunamadı"

        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active

        if ws.max_row <= 1:
            return "📊 Excel'de henüz veri yok"

        winning_patterns = {
            'high_cards': {'wins': 0, 'total': 0},
            'low_cards': {'wins': 0, 'total': 0},
            'mixed_suits': {'wins': 0, 'total': 0},
            'same_suits': {'wins': 0, 'total': 0}
        }

        for row in range(2, ws.max_row + 1):
            player_cards = ws.cell(row=row, column=4).value
            banker_cards = ws.cell(row=row, column=5).value
            result = ws.cell(row=row, column=13).value

            if player_cards and banker_cards and result in ['KAZANÇ', 'KAYIP']:
                high_cards = re.findall(r'(10|J|Q|K|A)[♠♥♦♣]', player_cards + banker_cards)
                if len(high_cards) >= 3:
                    winning_patterns['high_cards']['total'] += 1
                    if result == 'KAZANÇ':
                        winning_patterns['high_cards']['wins'] += 1

                low_cards = re.findall(r'([2-7])[♠♥♦♣]', player_cards + banker_cards)
                if len(low_cards) >= 4:
                    winning_patterns['low_cards']['total'] += 1
                    if result == 'KAZANÇ':
                        winning_patterns['low_cards']['wins'] += 1

                suits = re.findall(r'[♠♥♦♣]', player_cards + banker_cards)
                unique_suits = len(set(suits))
                if unique_suits >= 3:
                    winning_patterns['mixed_suits']['total'] += 1
                    if result == 'KAZANÇ':
                        winning_patterns['mixed_suits']['wins'] += 1
                elif unique_suits == 1:
                    winning_patterns['same_suits']['total'] += 1
                    if result == 'KAZANÇ':
                        winning_patterns['same_suits']['wins'] += 1

        return generate_winning_patterns_report(winning_patterns)

    except Exception as e:
        return f"❌ Kazanan kart desenleri analiz hatası: {e}"

def generate_winning_patterns_report(winning_patterns):
    report = "💎 **KAZANAN KART DESENLERİ ANALİZİ** 💎\n\n"

    for pattern, stats in winning_patterns.items():
        if stats['total'] > 0:
            win_rate = (stats['wins'] / stats['total']) * 100
            pattern_name = get_pattern_display_name(pattern)
            report += f"🎯 {pattern_name}:\n"
            report += f"   📊 {stats['wins']}/{stats['total']} (%{win_rate:.1f} başarı)\n\n"

    return report

def get_pattern_display_name(pattern):
    names = {
        'high_cards': '🔼 YÜKSEK KARTLAR (8-9-10-J-Q-K-A)',
        'low_cards': '🔻 DÜŞÜK KARTLAR (2-3-4-5-6-7)',
        'mixed_suits': '🌈 KARIŞIK RENKLER',
        'same_suits': '🎯 TEK RENK'
    }
    return names.get(pattern, pattern)

def analyze_c2_specific_strategies():
    try:
        if not os.path.exists(EXCEL_FILE):
            return "❌ Excel dosyası bulunamadı"

        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active

        if ws.max_row <= 1:
            return "📊 Excel'de henüz veri yok"

        c2_strategies = {}

        for row in range(2, ws.max_row + 1):
            c2_type = ws.cell(row=row, column=9).value
            player_cards = ws.cell(row=row, column=4).value
            banker_cards = ws.cell(row=row, column=5).value
            result = ws.cell(row=row, column=13).value
            color_prediction = ws.cell(row=row, column=10).value

            if c2_type and c2_type != 'N/A' and player_cards and banker_cards:
                if c2_type not in c2_strategies:
                    c2_strategies[c2_type] = {
                        'total_games': 0,
                        'wins': 0,
                        'successful_colors': defaultdict(int),
                        'total_colors': defaultdict(int)
                    }

                c2_strategies[c2_type]['total_games'] += 1
                if result == 'KAZANÇ':
                    c2_strategies[c2_type]['wins'] += 1

                if color_prediction:
                    for color in ['🔴', '🔵', '🟢', '⚫']:
                        if color in color_prediction:
                            c2_strategies[c2_type]['total_colors'][color] += 1
                            if result == 'KAZANÇ':
                                c2_strategies[c2_type]['successful_colors'][color] += 1

        return generate_c2_strategies_report(c2_strategies)

    except Exception as e:
        return f"❌ C2 stratejileri analiz hatası: {e}"

def generate_c2_strategies_report(c2_strategies):
    if not c2_strategies:
        return "📊 C2 stratejisi verisi bulunamadı"

    report = "🎯 **C2 TİPLERİNE ÖZEL STRATEJİLER** 🎯\n\n"

    for c2_type, stats in c2_strategies.items():
        total_games = stats['total_games']
        win_rate = (stats['wins'] / total_games * 100) if total_games > 0 else 0
        
        report += f"🔹 **{c2_type}** (%{win_rate:.1f} başarı):\n"

        best_color = None
        best_rate = 0
        
        for color in ['🔴', '🔵', '🟢', '⚫']:
            total_color = stats['total_colors'][color]
            successful_color = stats['successful_colors'][color]
            
            if total_color > 0:
                color_win_rate = (successful_color / total_color) * 100
                report += f"   {color}: %{color_win_rate:.1f} ({successful_color}/{total_color})\n"
                
                if color_win_rate > best_rate:
                    best_rate = color_win_rate
                    best_color = color

        if best_color:
            report += f"   ✅ **ÖNERİ:** {best_color} odaklan\n"

        report += "\n" + "─" * 40 + "\n\n"

    return report

# EXCEL TABANLI OTOMATİK SİSTEM
async def excel_based_sistemi(game_info):
    print("📊 EXCEL TABANLI SİSTEM analiz başlıyor...")
    
    try:
        recommended_c2, c2_reason = get_c2_recommendation_from_excel()
        print(f"🎯 Excel Tavsiyesi: {recommended_c2} - {c2_reason}")
        
        smart_color, smart_reason = predict_color_by_c2_excel(recommended_c2, game_info['player_cards'])
        print(f"🎨 Akıllı Renk: {smart_color} - {smart_reason}")
        
        risk_seviye, risk_uyarilar = super_risk_analizi()
        
        filtre_gecen = 0
        toplam_filtre = 6
        
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        if ws.max_row >= 20:
            filtre_gecen += 1
        else:
            print("🚫 Excel Sistemi: Yeterli veri yok")
            return
        
        c2_stats = get_c2_stats_from_excel(recommended_c2)
        if c2_stats and c2_stats['total'] >= 10 and c2_stats['win_rate'] >= 60:
            filtre_gecen += 1
        else:
            print(f"🚫 Excel Sistemi: {recommended_c2} yeterli güvende değil")
            return
        
        if risk_seviye != "🔴 YÜKSEK RİSK":
            filtre_gecen += 1
        else:
            print(f"🚫 Excel Sistemi: Yüksek risk - {risk_uyarilar}")
            return
        
        daily = get_daily_stats()
        if daily['profit'] >= -15:
            filtre_gecen += 1
        else:
            print("🚫 Excel Sistemi: Günlük performans düşük")
            return
        
        if len(color_trend) >= 5:
            if color_trend[-5:].count(smart_color[0]) >= 1:
                filtre_gecen += 1
            else:
                print("🚫 Excel Sistemi: Trend desteklemiyor")
                return
        else:
            filtre_gecen += 1
        
        son_30_dk = datetime.now(GMT3) - timedelta(minutes=30)
        son_sinyaller = [s for s in performance_stats['signal_history'] 
                        if s['timestamp'] >= son_30_dk]
        if len(son_sinyaller) <= 5:
            filtre_gecen += 1
        else:
            print("🚫 Excel Sistemi: Son 30 dakikada çok fazla sinyal")
            return
        
        print(f"✅ Excel Filtreler: {filtre_gecen}/{toplam_filtre} geçti")
        
        if filtre_gecen < 4:
            print(f"🚫 Excel Sistemi: Yetersiz filtre ({filtre_gecen}/{toplam_filtre})")
            return
        
        next_game_num = get_next_game_number(game_info['game_number'])
        c2_3_info = {
            'c2_3_type': recommended_c2,
            'c2_3_description': f"EXCEL TAVSİYE - {c2_reason}"
        }
        
        await send_new_signal(next_game_num, smart_color[0],
                            f"📊 EXCEL SİSTEM - {smart_reason} | {filtre_gecen}/{toplam_filtre} Filtre", 
                            c2_3_info)
        print(f"🎯 EXCEL SİSTEM sinyal gönderildi: #{next_game_num}")
        
    except Exception as e:
        print(f"❌ Excel tabanlı sistem hatası: {e}")

def get_c2_stats_from_excel(c2_type):
    try:
        if not os.path.exists(EXCEL_FILE):
            return None
            
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        
        wins, total = 0, 0
        for row in range(2, ws.max_row + 1):
            row_c2_type = ws.cell(row=row, column=9).value
            result = ws.cell(row=row, column=13).value
            
            if row_c2_type == c2_type:
                total += 1
                if result == 'KAZANÇ':
                    wins += 1
        
        win_rate = (wins / total * 100) if total > 0 else 0
        return {'total': total, 'wins': wins, 'win_rate': win_rate}
        
    except Exception as e:
        print(f"❌ C2 istatistik hatası: {e}")
        return None

# YENİ KOMUT: EXCEL GÖNDER
@client.on(events.NewMessage(pattern='(?i)/excel_gonder'))
async def handle_excel_gonder(event):
    """Excel dosyasını Telegram'a gönderir"""
    try:
        if event.sender_id != ADMIN_ID:
            await event.reply("❌ Bu komut sadece admin için!")
            return

        if not os.path.exists(EXCEL_FILE):
            await event.reply("❌ Excel dosyası bulunamadı!")
            return

        # Dosya boyutu kontrolü
        file_size = os.path.getsize(EXCEL_FILE)
        file_size_mb = file_size / (1024 * 1024)
        
        if file_size_mb > 50:
            await event.reply(f"❌ Excel dosyası çok büyük ({file_size_mb:.1f}MB). 50MB altındaki dosyalar gönderilebilir.")
            return

        # Excel istatistiklerini hazırla
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        total_records = ws.max_row - 1
        
        # Son 10 kaydı al
        recent_games_info = ""
        for row in range(max(2, ws.max_row - 9), ws.max_row + 1):
            game_num = ws.cell(row=row, column=1).value
            result = ws.cell(row=row, column=13).value
            color_pred = ws.cell(row=row, column=10).value
            if game_num and result:
                recent_games_info += f"#{game_num} {result} {color_pred}\n"

        await event.reply(f"📊 **Excel Dosyası Gönderiliyor...**\n\n"
                         f"📁 Dosya: `{EXCEL_FILE}`\n"
                         f"📈 Toplam Kayıt: {total_records}\n"
                         f"💾 Boyut: {file_size_mb:.1f}MB\n\n"
                         f"🎯 **Son 10 Kayıt:**\n{recent_games_info}")

        # Dosyayı gönder
        await client.send_file(
            event.chat_id,
            EXCEL_FILE,
            caption=f"📈 **Royal Baccarat Veri Tablosu**\n\n"
                   f"📊 Toplam {total_records} kayıt\n"
                   f"🕒 Son güncelleme: {datetime.now(GMT3).strftime('%d.%m.%Y %H:%M')}\n"
                   f"💾 Dosya boyutu: {file_size_mb:.1f}MB\n\n"
                   f"🔍 **İçerik:**\n"
                   f"• Tüm oyun geçmişi\n"
                   f"• C2 tipi analizleri\n"
                   f"• Renk tahminleri\n"
                   f"• Performans istatistikleri"
        )
        
        print(f"✅ Excel dosyası {event.chat_id} numaralı sohbete gönderildi")

    except Exception as e:
        error_msg = f"❌ Excel gönderme hatası: {e}"
        await event.reply(error_msg)
        print(error_msg)

# DİĞER FONKSİYONLAR (kısaltma nedeniyle özet)
def get_suit_display_name(suit_symbol):
    suit_names = {'♠': '♠️ MAÇA', '♥': '❤️ KALP', '♦': '♦️ KARO', '♣': '♣️ SİNEK'}
    return suit_names.get(suit_symbol, f"❓ {suit_symbol}")

def get_baccarat_value(card_char):
    if card_char == '10': return 10
    if card_char in 'AKQJ2T': return 0
    elif card_char.isdigit(): return int(card_char)
    return -1

def extract_largest_value_suit(cards_str):
    try:
        cards = re.findall(r'(10|[A2-9TJQK])([♣♦♥♠])', cards_str)
        if not cards: return None
        max_value, largest_value_suit = -1, None
        values = [get_baccarat_value(card[0]) for card in cards]
        if len(values) == 2 and values[0] == values[1]: return None
        for card_char, suit in cards:
            value = get_baccarat_value(card_char)
            if value > max_value: max_value, largest_value_suit = value, suit
        return largest_value_suit if max_value > 0 else None
    except Exception as e:
        print(f"❌ extract_largest_value_suit hatası: {e}")
        return None

def analyze_simple_pattern(player_cards, banker_cards, game_number):
    try:
        signal_color = extract_largest_value_suit(player_cards)
        if not signal_color: return None, "Renk tespit edilemedi"
        color_trend.append(signal_color)
        if len(color_trend) > 10: color_trend.pop(0)
        player_card_data = re.findall(r'(10|[A2-9TJQK])([♣♦♥♠])', player_cards)
        player_values = [get_baccarat_value(card[0]) for card in player_card_data]
        banker_card_data = re.findall(r'(10|[A2-9TJQK])([♣♦♥♠])', banker_cards)
        banker_values = [get_baccarat_value(card[0]) for card in banker_card_data]
        total_cards = len(player_values) + len(banker_values)
        if sum(player_values) >= 8 and len(player_values) >= 3: return signal_color, "🎯 GÜÇLÜ EL"
        elif sum(player_values) in [8, 9]: return signal_color, "🏆 DOĞAL KAZANÇ"
        elif total_cards >= 5: return signal_color, "📊 5+ KART"
        elif len(color_trend) >= 3 and color_trend[-3:] == [signal_color] * 3: return signal_color, "🚨 3x TEKRAR"
        else: return signal_color, "📈 STANDART SİNYAL"
    except Exception as e:
        print(f"❌ Pattern analysis error: {e}")
        return None, f"Hata: {e}"

# [DİĞER FONKSİYONLAR AYNI ŞEKİLDE DEVAM EDER...]
# predict_color_by_c2_excel, get_c2_recommendation_from_excel, vs. tüm fonksiyonlar

# KART ANALİZ KOMUTLARI
@client.on(events.NewMessage(pattern='(?i)/kart_analiz'))
async def handle_kart_analiz(event):
    analysis = analyze_cards_by_c2_type()
    await event.reply(analysis)

@client.on(events.NewMessage(pattern='(?i)/kart_desen'))
async def handle_kart_desen(event):
    analysis = analyze_winning_cards_patterns()
    await event.reply(analysis)

@client.on(events.NewMessage(pattern='(?i)/c2_strateji'))
async def handle_c2_strateji(event):
    analysis = analyze_c2_specific_strategies()
    await event.reply(analysis)

@client.on(events.NewMessage(pattern='(?i)/excel_tahmin'))
async def handle_excel_tahmin(event):
    try:
        if not game_results:
            return await event.reply("📊 Henüz oyun verisi yok!")
        
        recommended_c2, c2_reason = get_c2_recommendation_from_excel()
        last_game_num = max(game_results.keys())
        last_game = game_results[last_game_num]
        
        smart_color, smart_reason = predict_color_by_c2_excel(recommended_c2, last_game['player_cards'])
        risk_seviye, risk_uyarilar = super_risk_analizi()
        c2_stats = get_c2_stats_from_excel(recommended_c2)
        
        tahmin_mesaji = f"""🎯 **EXCEL TABANLI TAHMİN** 🎯

📊 **Excel Analizi:**
• Önerilen C2 Tipi: {recommended_c2}
• Sebep: {c2_reason}
• İstatistik: {c2_stats['total']} oyun, %{c2_stats['win_rate']:.1f} başarı

🎨 **TAHMİN:**
• Renk: {smart_color}
• Sebep: {smart_reason}

⚡ **RİSK ANALİZİ:**
• Seviye: {risk_seviye}
• Uyarılar: {', '.join(risk_uyarilar) if risk_uyarilar else 'Yok'}

💡 **STRATEJİ:**
{generate_excel_strategy(recommended_c2, smart_color, c2_stats)}

📈 **GÜVEN:** {calculate_confidence_level(c2_stats, risk_seviye)}
"""
        await event.reply(tahmin_mesaji)
        
    except Exception as e:
        await event.reply(f"❌ Excel tahmin hatası: {e}")

def generate_excel_strategy(c2_type, color_prediction, c2_stats):
    strategy = ""
    
    if c2_type == '#C2_3':
        strategy += "🎯 **KLASİK STRATEJİ:** Excel verilerine göre en güvenilir tip. "
    elif c2_type == '#C2_2':
        strategy += "🔄 **ALTERNATİF STRATEJİ:** Dengeli yaklaşım önerilir. "
    elif c2_type == '#C3_2':
        strategy += "⚡ **VARYANT STRATEJİ:** Dikkatli ilerleyin. "
    elif c2_type == '#C3_3':
        strategy += "🌟 **ÖZEL STRATEJİ:** Nadir pattern, özel taktik. "
    
    if c2_stats['win_rate'] >= 75:
        strategy += "✅ **YÜKSEK GÜVEN:** Büyük bahisler düşünülebilir."
    elif c2_stats['win_rate'] >= 65:
        strategy += "📊 **ORTA GÜVEN:** Standart bahislerle devam edin."
    else:
        strategy += "⚠️ **DÜŞÜK GÜVEN:** Küçük bahislerle test edin."
    
    color_emoji = color_prediction.split()[0] if isinstance(color_prediction, str) else color_prediction[0]
    if color_emoji == '🔴':
        strategy += "\n🔴 **KIRMIZI TAKTİK:** Agresif oynanabilir."
    elif color_emoji == '🔵':
        strategy += "\n🔵 **MAVİ TAKTİK:** Dengeli ve sabit kal."
    elif color_emoji == '🟢':
        strategy += "\n🟢 **YEŞİL TAKTİK:** Riskleri iyi yönet."
    
    return strategy

def calculate_confidence_level(c2_stats, risk_seviye):
    base_confidence = min(100, c2_stats['win_rate'] * 1.2)
    
    if risk_seviye == "🔴 YÜKSEK RİSK":
        base_confidence *= 0.7
    elif risk_seviye == "🟡 ORTA RİSK":
        base_confidence *= 0.85
    
    if c2_stats['total'] < 10:
        base_confidence *= 0.8
    elif c2_stats['total'] < 20:
        base_confidence *= 0.9
    
    return f"%{base_confidence:.1f}"

# YENİ MOD KOMUTU
@client.on(events.NewMessage(pattern='(?i)/mod_excel'))
async def handle_mod_excel(event):
    global SISTEM_MODU
    SISTEM_MODU = "excel_based"
    
    excel_status = "✅ Excel mevcut" if os.path.exists(EXCEL_FILE) else "❌ Excel bulunamadı"
    
    record_count = 0
    if os.path.exists(EXCEL_FILE):
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active
            record_count = ws.max_row - 1
        except:
            record_count = 0
    
    await event.reply(f"""📊 **EXCEL TABANLI MOD AKTİF!**

{excel_status}
📈 Kayıtlı Veri: {record_count} oyun

🎯 **ÖZELLİKLER:**
• Excel'deki tüm geçmiş verileri analiz eder
• En başarılı C2 tipini otomatik seçer
• Risk analizi ile filtre uygular
• Gerçek istatistiklere dayalı tahmin yapar

⚡ **ÇALIŞMA MANTIĞI:**
1. Excel'den en başarılı C2 tipini seçer
2. Bu tip için en kazançlı rengi belirler  
3. 6 farklı filtre uygular
4. Güvenilir sinyalleri otomatik gönderir

💡 **TAVSİYE:** {record_count}+ kayıt sonrası maksimum verim alınır!""")

# ANA MESAJ İŞLEYİCİ
@client.on(events.NewMessage(chats=KANAL_KAYNAK_ID))
@client.on(events.MessageEdited(chats=KANAL_KAYNAK_ID))
async def handle_source_channel_message(event):
    try:
        message = event.message
        text = message.text or ""
        cleaned_text = re.sub(r'\*\*', '', text).strip()
        cleaned_text = re.sub(r'\s+', ' ', cleaned_text).strip()
        gmt3_time = datetime.now(GMT3).strftime('%H:%M:%S')
        print(f"[{gmt3_time}] 📥 Mesaj: #{len(cleaned_text)} karakter")
        game_info = extract_game_info_from_message(cleaned_text)
        if game_info['game_number'] is None: return
        game_results[game_info['game_number']] = game_info
        
        await save_game_result_to_excel(game_info)
        
        await check_martingale_trackers()
        if not is_signal_active:
            if game_info['is_final'] and game_info.get('is_c2_3'):
                trigger_game_num, c2_3_type, c2_3_desc = game_info['game_number'], game_info['c2_3_type'], game_info['c2_3_description']
                print(f"🎯 {c2_3_desc} tespit edildi: #{trigger_game_num} - {c2_3_type}")
                
                if SISTEM_MODU == "normal_hibrit": 
                    await normal_hibrit_sistemi(game_info)
                elif SISTEM_MODU == "super_hibrit": 
                    await super_hibrit_sistemi(game_info)
                elif SISTEM_MODU == "quantum_hibrit":
                    await quantum_hibrit_sistemi(game_info)
                elif SISTEM_MODU == "quantum_pro":
                    await quantum_pro_sistemi(game_info)
                elif SISTEM_MODU == "master_elite":
                    await master_elite_sistemi(game_info)
                elif SISTEM_MODU == "excel_based":
                    await excel_based_sistemi(game_info)
                    
    except Exception as e: 
        print(f"❌ Mesaj işleme hatası: {e}")

# YARDIM MENÜSÜNÜ GÜNCELLE
@client.on(events.NewMessage(pattern='(?i)/yardim'))
async def handle_yardim(event):
    yardim_mesaji = """🤖 **ROYAL BACCARAT BOT - YARDIM MENÜSÜ** 🤖

🎯 **TEMEL KOMUTLAR:**
• /basla - Botu başlat
• /durum - Sistem durumu
• /istatistik - Detaylı istatistikler
• /performans - Performans raporu
• /rapor - Günlük/haftalık rapor

📊 **ANALİZ KOMUTLARI:**
• /c2analiz - C2-3 tip performansları
• /pattern - Pattern performans tablosu
• /trend - Trend analizi
• /eniyi - En iyi performans
• /enkotu - En kötü performans
• /tavsiye - Trading tavsiyesi

💾 **EXCEL & C2 ANALİZ:**
• /excel - Excel dosyası durumu
• /excel_analiz - Excel veri analizi
• /excel_c2 - Excel C2 tip analizi
• /c2_tavsiye - Excel C2 tavsiyesi
• /renk_tahmini - Renk tahmin analizi
• /smart_renk - Akıllı renk tahmini
• /excel_tahmin - Excel tabanlı tahmin
• /excel_gonder - 📁 Excel dosyasını gönder (ADMIN)

🃏 **KART ANALİZ KOMUTLARI:**
• /kart_analiz - C2 tiplerine göre kart analizi
• /kart_desen - Kazanan kart desenleri
• /c2_strateji - C2 özel stratejiler
• /tum_analiz - Tüm analizler

🎛️ **SİSTEM MODLARI:**
• /mod_normal - Normal Hibrit Mod
• /mod_super - Süper Hibrit Mod  
• /mod_quantum - Quantum Hibrit Mod
• /mod_quantumpro - Quantum Pro Mod
• /mod_masterelite - Master Elite Mod
• /mod_excel - 📊 EXCEL TABANLI MOD (YENİ!)
• /mod_durum - Aktif modu göster

⚡ **ADMIN KOMUTLARI:**
• /temizle - Trend verilerini temizle
• /acil_durdur - Acil durdurma
• /aktif_et - Sistemi tekrar aktif et
• /excel_gonder - Excel dosyasını gönder

🔧 **Sistem:** Hibrit Pattern + Martingale {MAX_MARTINGALE_STEPS} Seviye
🕒 **Saat Dilimi:** GMT+3 (İstanbul)
💾 **Excel Kayıt:** Tüm veriler otomatik kaydedilir
🎯 **Excel Mod:** Gerçek verilere dayalı akıllı sinyaller
""".format(MAX_MARTINGALE_STEPS=MAX_MARTINGALE_STEPS)
    await event.reply(yardim_mesaji)

# CLIENT TANIMLAMA
client = TelegramClient('kolera_bot', API_ID, API_HASH)

if __name__ == '__main__':
    print("🤖 ROYAL BACCARAT BOT BAŞLIYOR...")
    print(f"🔧 API_ID: {API_ID}")
    print(f"🎯 Kaynak Kanal: {KANAL_KAYNAK_ID}")
    print(f"📤 Hedef Kanal: {KANAL_HEDEF}")
    print(f"👤 Admin ID: {ADMIN_ID}")
    print(f"🎛️ Varsayılan Mod: {SISTEM_MODU}")
    print(f"💾 Excel Kayıt Sistemi: AKTİF")
    print(f"📊 C2-3 Analiz Sistemi: AKTİF")
    print(f"📈 Pattern Performans Takibi: AKTİF")
    print(f"🎨 Renk Oyunu Tahmini: AKTİF")
    print(f"🤖 Excel C2 Analizi: AKTİF")
    print(f"🃏 Kart Analiz Sistemi: AKTİF")
    print(f"📤 Excel Gönderme: AKTİF")
    print(f"⚛️ Quantum Hibrit Sistem: AKTİF")
    print(f"🚀 Quantum PRO Sistem: AKTİF")
    print(f"🏆 Master Elite Sistem: AKTİF")
    print(f"🕒 Saat Dilimi: GMT+3")
    
    init_excel_file()
    
    print("⏳ Bağlanıyor...")
    try:
        with client: 
            client.start(bot_token=BOT_TOKEN)
            client.run_until_disconnected()
    except KeyboardInterrupt: 
        print("\n👋 Bot durduruluyor...")
    except Exception as e: 
        print(f"❌ Bot başlangıç hatası: {e}")
